import io
import hashlib
import json
import re
import zipfile
import xml.etree.ElementTree as ET
from copy import deepcopy
from datetime import datetime, time, timedelta, timezone
from difflib import SequenceMatcher
from html import escape
from pathlib import Path
from typing import Iterable, List
from urllib.parse import quote
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, EmailStr
from sqlalchemy import delete, func, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_user_or_dev_user, get_db
from app.core.errors import ErrorCode, build_error
from app.core.response import success_response
from app.core.transnews_client import TransNewsClient
from app.models.article import Article
from app.models.article_analysis import ArticleAnalysis
from app.models.article_match import ArticleMatch
from app.models.crawl_run import CrawlRun
from app.models.email_delivery import EmailDelivery
from app.models.dongguk_article_trash import DonggukArticleTrash
from app.models.dongguk_preview_cache import DonggukPreviewCache
from app.models.dongguk_mail_draft import DonggukMailDraft
from app.models.importance_score import ImportanceScore
from app.models.keyword import Keyword
from app.models.social_metric import SocialMetric
from app.models.summary import Summary
from app.models.user import User
import app.schemas.articles
from app.services.article_service import ArticleService
from app.services.article_identity import (
    canonicalize_article_url,
    is_same_publisher_article,
    normalize_article_title,
)
from app.services.dify_service import DifyService
from app.services.email_service import EmailService
from app.services.holiday_service import HolidayService, article_collection_window
from app.services.priority_insight_service import PriorityInsightService

router = APIRouter(prefix="/reports", tags=["reports"])

BLUE = "3B82F6"
BLUE_DARK = "1E3A8A"
BLUE_SOFT = "EFF6FF"
LINE = "DBEAFE"
HEADER_FILL = "2563EB"
TEXT = "172033"
MUTED = "64748B"
GREEN = "16A34A"
RED = "DC2626"
AMBER = "D97706"
KST = ZoneInfo("Asia/Seoul")
DONGGUK_MAIL_POLICY_VERSION = 11
DEFAULT_DONGGUK_PRIORITY_CRITERIA = """홍보처 AI 기사 선정 기준

우선순위 기준:
- 총장 또는 이사장의 공식 메시지가 포함된 기사를 가장 먼저 선정합니다.
- 기부, 장학금, 발전기금처럼 학교 이미지와 직접 연결되는 기사를 우선 선정합니다.
- 건학 120주년 등 현재 진행 중인 학교 캠페인 관련 기사를 우선 선정합니다.
- 동국대 공식 홈페이지(dongguk.edu) 게시글, 공식 보도자료, 공식 행사, Dream Workshop, C포럼, 서울국제명상엑스포 등 홍보처가 직접 배포하는 성격의 기사를 우선 선정합니다.
- 동문소식은 [동문소식] 형식이거나 동문 이름/학번/학과가 명확하고 사회적 성취가 드러나는 경우 우선 선정합니다.
- 연구 성과, 기술 개발, 특허, AI 관련 성과 기사를 우선 선정합니다.
- 공식 기관과의 협약, 사업 선정, 컨소시엄 참여 기사를 우선 선정합니다.
- 학교, 교직원, 학생의 수상 및 공식 인증 획득 기사를 우선 선정합니다.
- 학교가 주최하거나 공식적으로 참여한 주요 행사 기사를 우선 선정합니다.
- 학술대회, 세미나, 입시, 교육 프로그램, 인사, 동문 및 교수 인터뷰 기사는 직접적인 학교 성과 기사보다 낮게 선정합니다.
- 대학 [교육] 섹션은 개별 대학 홍보 기사보다 교육부, 대교협, 사총협, 고등교육법, 교육교부금, 등록금, 교원창업, AI 기본역량 등 고등교육 정책성 기사를 우선 선정합니다.
- 불교 [종단] 섹션은 조계종, 종단, 포교, 출가, 성보, 불교문화유산, 불교중앙박물관, 불교신문/법보신문/BBS/BTN/현대불교 등 종단성과 불교계 파급력이 있는 기사를 우선 선정합니다.

대표 기사 선정 기준:
- 동일 주제, 동일 보도자료, 같은 사건의 반복 보도는 하나의 그룹으로 묶고 대표 기사 1건만 선정합니다.
- 원문 URL이 정상이고 본문 전체를 확인할 수 있는 기사를 우선합니다.
- 기관명, 행사명, 인물명, 금액, 성과 등 핵심 정보가 제목에 명확히 드러난 기사를 우선합니다.
- 요약에 필요한 사실 정보가 충분하고 기사 내용이 충실한 기사를 우선합니다.
- 언론사 신뢰도와 홍보처 배포 활용도가 높은 기사를 우선합니다.

제외 기준:
- 동국대학교와 직접 관련성이 확인되지 않는 기사는 제외합니다.
- 원문 확인이 어렵거나 본문 정보가 부족한 기사는 제외합니다."""
DEFAULT_DONGGUK_REPRESENTATIVE_CRITERIA = """대표 기사 선정 기준:
- 동일 주제/동일 보도자료/같은 사건의 반복 보도는 하나의 그룹으로 묶고, 대표 기사 1건만 선정한다.
- 대표 기사는 원문 URL이 정상이고 본문 확인이 가능한 기사를 우선한다.
- 제목이 가장 명확하고 기관명, 행사명, 인물명, 금액, 성과 등 핵심 정보가 잘 드러난 기사를 우선한다.
- 기사 내용이 길고 요약에 필요한 사실 정보가 충분한 기사를 우선한다.
- 출처 신뢰도와 홍보처 배포 활용성을 고려한다.
- 단순 재전송, 제목만 바꾼 기사, 내용이 짧거나 원문 확인이 어려운 기사는 제외한다."""


def _normalize_dongguk_priority_criteria(criteria: str | None) -> str:
    text = (criteria or "").strip()
    if not text:
        return DEFAULT_DONGGUK_PRIORITY_CRITERIA
    return text
DONGGUK_DASHBOARD_URL = "https://2c25-210-94-172-73.ngrok-free.app/"
DONGGUK_HWPX_TEMPLATE = Path(__file__).resolve().parents[2] / "templates" / "dongguk_daily_news_template.hwpx"
HWPX_NS = {
    "hp": "http://www.hancom.co.kr/hwpml/2011/paragraph",
    "hs": "http://www.hancom.co.kr/hwpml/2011/section",
}

ARTICLE_HEADERS = [
    "기사 ID",
    "제목",
    "출처",
    "게시 일시",
    "키워드",
    "AI 감성",
    "홍보성",
    "중요도",
    "중요도 사유",
    "AI 요약",
    "원문 URL",
    "언어",
    "수집 일시",
]
ARTICLE_WIDTHS = [10, 52, 18, 18, 22, 12, 10, 10, 44, 62, 52, 8, 18]


class EmailReportRequest(BaseModel):
    to_emails: List[EmailStr]
    keyword_id: int | None = None
    keyword_name: str | None = None


class DonggukMailArticle(BaseModel):
    id: int | None = None
    title: str
    source: str | None = None
    collection_source: str | None = None
    section: str | None = None
    category: str = "기타"
    summary: str | None = None
    url: str | None = None
    thumbnail_url: str | None = None
    published_at: str | None = None
    links: list[str] = []
    priority: str | None = None
    priority_name: str | None = None
    score: int | float | None = None
    is_syndicated: bool = False
    selection_reason: str | None = None


class DonggukEmailRequest(BaseModel):
    to_emails: List[EmailStr]
    subject: str
    articles: list[DonggukMailArticle]
    keyword_id: int | None = None
    mail_date: str | None = None
    exclude_similar_sent: bool = True
    use_current_articles: bool = False
    priority_criteria: str | None = None
    is_test: bool = False


class DonggukPreviewRequest(BaseModel):
    subject: str
    articles: list[DonggukMailArticle]
    keyword_id: int | None = None
    mail_date: str | None = None
    exclude_similar_sent: bool = True
    force_rebuild: bool = False
    selected_article_keys: list[str] = []
    removed_article_keys: list[str] = []
    removed_articles: list[DonggukMailArticle] = []
    priority_criteria: str | None = None


class DonggukDraftRequest(BaseModel):
    subject: str
    keyword_id: int | None = None
    mail_date: str
    selected_article_keys: list[str] = []
    selected_articles: list[DonggukMailArticle] = []
    removed_article_keys: list[str] = []
    removed_articles: list[DonggukMailArticle] = []
    preview_data: dict | None = None
    feedback_source: str | None = None


class DonggukTrashRequest(BaseModel):
    keyword_id: int | None = None
    mail_date: str
    article: DonggukMailArticle


class DonggukTrashActionRequest(BaseModel):
    keyword_id: int | None = None
    mail_date: str
    article_id: int


class DonggukHwpRequest(BaseModel):
    subject: str
    articles: list[DonggukMailArticle]
    mail_date: str | None = None


class DonggukLinkPreviewRequest(BaseModel):
    url: str


def _fmt_dt(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def _score_value(row) -> float | None:
    score = row["score"]
    return float(score) if score is not None else None


def _sentiment_key(value: str | None) -> str:
    text = str(value or "").lower()
    if "부정" in text or "negative" in text:
        return "부정"
    if "긍정" in text or "positive" in text:
        return "긍정"
    if "중립" in text or "neutral" in text:
        return "중립"
    return "미분류"


def _promotion_label(value) -> str:
    if value is True:
        return "홍보성"
    if value is False:
        return "일반"
    return "미분류"


async def _keyword_name(db: AsyncSession, current_user: User, keyword_id: int | None) -> str | None:
    if not keyword_id:
        return None
    result = await db.execute(
        select(Keyword.keyword_text).where(
            Keyword.id == keyword_id,
            Keyword.user_id == current_user.id,
        )
    )
    return result.scalar_one_or_none()


def _build_query(current_user: User, keyword_id: int | None = None):
    latest_summary_subq = (
        select(
            Summary.article_id,
            Summary.summary_text,
            func.row_number()
            .over(partition_by=Summary.article_id, order_by=Summary.created_at.desc())
            .label("rn"),
        ).subquery()
    )

    latest_importance_subq = (
        select(
            ImportanceScore.article_id,
            ImportanceScore.score,
            ImportanceScore.reason.label("importance_reason"),
            func.row_number()
            .over(
                partition_by=ImportanceScore.article_id,
                order_by=ImportanceScore.created_at.desc(),
            )
            .label("rn"),
        )
        .where(ImportanceScore.user_id == current_user.id)
        .where(ImportanceScore.is_current.is_(True))
        .subquery()
    )

    keyword_subq = (
        select(
            ArticleMatch.article_id,
            func.string_agg(Keyword.keyword_text, ", ").label("keywords"),
        )
        .join(Keyword, Keyword.id == ArticleMatch.keyword_id)
        .where(Keyword.user_id == current_user.id)
        .group_by(ArticleMatch.article_id)
        .subquery()
    )

    stmt = (
        select(
            Article.id,
            Article.title,
            Article.publisher,
            Article.published_at,
            Article.url,
            Article.thumbnail_url,
            Article.language,
            Article.created_at,
            keyword_subq.c.keywords,
            ArticleAnalysis.sentiment,
            ArticleAnalysis.is_promotion,
            latest_importance_subq.c.score,
            latest_importance_subq.c.importance_reason,
            latest_summary_subq.c.summary_text,
        )
        .outerjoin(keyword_subq, keyword_subq.c.article_id == Article.id)
        .outerjoin(
            latest_importance_subq,
            (latest_importance_subq.c.article_id == Article.id)
            & (latest_importance_subq.c.rn == 1),
        )
        .outerjoin(
            latest_summary_subq,
            (latest_summary_subq.c.article_id == Article.id)
            & (latest_summary_subq.c.rn == 1),
        )
        .outerjoin(ArticleAnalysis, ArticleAnalysis.article_id == Article.id)
        .where(
            Article.id.in_(
                select(ArticleMatch.article_id)
                .join(Keyword, Keyword.id == ArticleMatch.keyword_id)
                .where(Keyword.user_id == current_user.id)
            )
        )
        .order_by(Article.published_at.desc().nullslast(), Article.id.desc())
    )

    if keyword_id:
        stmt = stmt.where(
            Article.id.in_(
                select(ArticleMatch.article_id)
                .join(Keyword, Keyword.id == ArticleMatch.keyword_id)
                .where(
                    ArticleMatch.keyword_id == keyword_id,
                    Keyword.user_id == current_user.id,
                )
            )
        )

    return stmt


async def _get_social_report_rows(db: AsyncSession, current_user: User, keyword_id: int | None = None) -> list[dict]:
    latest_subq = (
        select(
            SocialMetric.keyword_id,
            SocialMetric.source,
            func.max(SocialMetric.sampled_at).label("sampled_at"),
        )
        .where(SocialMetric.user_id == current_user.id)
        .group_by(SocialMetric.keyword_id, SocialMetric.source)
        .subquery()
    )
    stmt = (
        select(
            SocialMetric.keyword_text,
            SocialMetric.source,
            SocialMetric.mention_count,
            SocialMetric.positive_hint_count,
            SocialMetric.negative_hint_count,
            SocialMetric.sampled_at,
        )
        .join(
            latest_subq,
            (latest_subq.c.keyword_id == SocialMetric.keyword_id)
            & (latest_subq.c.source == SocialMetric.source)
            & (latest_subq.c.sampled_at == SocialMetric.sampled_at),
        )
        .where(SocialMetric.user_id == current_user.id)
    )
    if keyword_id:
        stmt = stmt.where(SocialMetric.keyword_id == keyword_id)
    result = await db.execute(stmt)
    return [dict(row) for row in result.mappings().all()]


def _report_context(rows: Iterable[dict], keyword_name: str | None = None, social_rows: list[dict] | None = None) -> dict:
    rows = list(rows)
    keyword_counts: dict[str, int] = {}
    sentiment_counts = {"긍정": 0, "중립": 0, "부정": 0, "미분류": 0}
    promotion_counts = {"홍보성": 0, "일반": 0, "미분류": 0}
    daily_counts: dict[str, int] = {}
    scores: list[float] = []
    summarized_count = 0
    social_source_counts: dict[str, int] = {}
    social_total = 0
    social_positive = 0
    social_negative = 0

    for row in rows:
        for keyword in (row["keywords"] or "미분류").split(","):
            key = keyword.strip() or "미분류"
            keyword_counts[key] = keyword_counts.get(key, 0) + 1

        sentiment_counts[_sentiment_key(row["sentiment"])] += 1
        promotion_counts[_promotion_label(row["is_promotion"])] += 1

        date_key = _fmt_dt(row["published_at"])[:10] or "날짜 없음"
        daily_counts[date_key] = daily_counts.get(date_key, 0) + 1

        score = _score_value(row)
        if score is not None:
            scores.append(score)
        if row["summary_text"]:
            summarized_count += 1

    for row in social_rows or []:
        source = row["source"] or "unknown"
        count = int(row["mention_count"] or 0)
        social_source_counts[source] = social_source_counts.get(source, 0) + count
        social_total += count
        social_positive += int(row["positive_hint_count"] or 0)
        social_negative += int(row["negative_hint_count"] or 0)

    top_articles = sorted(
        [row for row in rows if _score_value(row) is not None],
        key=lambda item: _score_value(item) or 0,
        reverse=True,
    )[:8]
    negative_rate = round((sentiment_counts["부정"] / len(rows)) * 100) if rows else 0
    promotion_rate = round((promotion_counts["홍보성"] / len(rows)) * 100) if rows else 0
    social_negative_rate = round((social_negative / social_total) * 100) if social_total else 0

    return {
        "keyword_name": keyword_name,
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "total_count": len(rows),
        "summary_count": summarized_count,
        "scored_count": len(scores),
        "avg_score": round(sum(scores) / len(scores), 1) if scores else None,
        "negative_rate": negative_rate,
        "promotion_rate": promotion_rate,
        "social_total": social_total,
        "social_positive": social_positive,
        "social_negative": social_negative,
        "social_negative_rate": social_negative_rate,
        "keyword_counts": sorted(keyword_counts.items(), key=lambda item: item[1], reverse=True)[:10],
        "sentiment_counts": [(k, v) for k, v in sentiment_counts.items() if v],
        "promotion_counts": [(k, v) for k, v in promotion_counts.items() if v],
        "daily_counts": sorted(daily_counts.items()),
        "social_source_counts": sorted(social_source_counts.items(), key=lambda item: item[1], reverse=True),
        "top_articles": top_articles,
        "executive_summary": _executive_summary(
            keyword_name=keyword_name,
            total_count=len(rows),
            negative_rate=negative_rate,
            promotion_rate=promotion_rate,
            social_total=social_total,
            social_negative_rate=social_negative_rate,
            top_articles=top_articles,
        ),
    }


def _executive_summary(
    *,
    keyword_name: str | None,
    total_count: int,
    negative_rate: int,
    promotion_rate: int,
    social_total: int,
    social_negative_rate: int,
    top_articles: list,
) -> list[str]:
    label = keyword_name or "전체 키워드"
    lines = [f"{label} 기준 수집 기사 {total_count}건을 분석했습니다."]
    if top_articles:
        lines.append(f"가장 먼저 확인할 기사는 '{top_articles[0]['title'] or '제목 없음'}'입니다.")
    if negative_rate >= 25:
        lines.append(f"부정 기사 비중이 {negative_rate}%로 높아 이슈 대응 메시지 점검이 필요합니다.")
    elif negative_rate:
        lines.append(f"부정 기사 비중은 {negative_rate}%로 관리 가능한 수준입니다.")
    else:
        lines.append("현재 뚜렷한 부정 기사 신호는 없습니다.")
    if promotion_rate >= 20:
        lines.append(f"홍보성 기사 비중은 {promotion_rate}%입니다. 자연 노출과 유료/홍보 노출을 분리해 보고하세요.")
    if social_total:
        lines.append(f"SNS 최근 7일 신호는 {social_total}건이며 부정 힌트 비중은 {social_negative_rate}%입니다.")
    return lines


def _style_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color=LINE)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_table(ws, start_row: int, start_col: int, title: str, headers: list[str], rows: list[tuple]) -> int:
    ws.cell(start_row, start_col, title).font = Font(bold=True, size=12, color=BLUE_DARK)
    header_row = start_row + 1
    for idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(header_row, idx, header)
        cell.fill = PatternFill("solid", fgColor=BLUE_SOFT)
        cell.font = Font(bold=True, color=BLUE_DARK)
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows or [("데이터 없음", 0)], start=header_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            ws.cell(row_idx, col_idx, value)
    return header_row + max(len(rows), 1)


def _add_bar_chart(ws, min_col: int, min_row: int, max_row: int, anchor: str, title: str) -> None:
    if max_row <= min_row:
        return
    chart = BarChart()
    chart.title = title
    chart.height = 7
    chart.width = 12
    data = Reference(ws, min_col=min_col + 1, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, anchor)


def _add_line_chart(ws, min_col: int, min_row: int, max_row: int, anchor: str, title: str) -> None:
    if max_row <= min_row:
        return
    chart = LineChart()
    chart.title = title
    chart.height = 7
    chart.width = 12
    data = Reference(ws, min_col=min_col + 1, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, anchor)


def _build_excel(rows, keyword_name: str | None = None, social_rows: list[dict] | None = None) -> Workbook:
    rows = list(rows)
    context = _report_context(rows, keyword_name, social_rows)
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "보고서 요약"
    summary_ws["A1"] = "AI 뉴스 인텔리전스 리포트"
    summary_ws["A1"].font = Font(bold=True, size=20, color=BLUE_DARK)
    summary_ws["A2"] = f"생성 시각: {context['generated_at']}"
    summary_ws["A3"] = f"키워드: {keyword_name or '전체 키워드'}"

    metrics = [
        ("전체 기사", context["total_count"]),
        ("AI 요약 완료", context["summary_count"]),
        ("중요도 산정", context["scored_count"]),
        ("평균 중요도", context["avg_score"] if context["avg_score"] is not None else "-"),
        ("부정 비중", f"{context['negative_rate']}%"),
        ("홍보성 비중", f"{context['promotion_rate']}%"),
        ("SNS 언급", context["social_total"]),
        ("SNS 부정 힌트", f"{context['social_negative_rate']}%"),
    ]
    for idx, (label, value) in enumerate(metrics, start=1):
        col = ((idx - 1) % 4) * 2 + 1
        row = 5 if idx <= 4 else 8
        summary_ws.cell(row, col, label).font = Font(bold=True, color=MUTED)
        summary_ws.cell(row + 1, col, value).font = Font(bold=True, size=16, color=BLUE)

    summary_ws["A11"] = "Executive Summary"
    summary_ws["A11"].font = Font(bold=True, size=13, color=BLUE_DARK)
    for row_idx, line in enumerate(context["executive_summary"], start=12):
        summary_ws.cell(row_idx, 1, f"- {line}")

    for col in range(1, 9):
        summary_ws.column_dimensions[summary_ws.cell(1, col).column_letter].width = 18

    chart_ws = wb.create_sheet("그래프")
    keyword_end = _write_table(chart_ws, 1, 1, "키워드별 기사 수", ["키워드", "기사 수"], context["keyword_counts"])
    sentiment_end = _write_table(chart_ws, 1, 5, "감성 분포", ["감성", "건수"], context["sentiment_counts"])
    promotion_end = _write_table(chart_ws, 16, 1, "홍보성 분포", ["구분", "건수"], context["promotion_counts"])
    daily_end = _write_table(chart_ws, 16, 5, "일자별 기사 추이", ["날짜", "기사 수"], context["daily_counts"][-30:])
    social_end = _write_table(chart_ws, 31, 1, "SNS 플랫폼별 언급", ["플랫폼", "언급 수"], context["social_source_counts"])
    _add_bar_chart(chart_ws, 1, 2, keyword_end, "H1", "키워드별 기사 수")
    _add_bar_chart(chart_ws, 5, 2, sentiment_end, "H16", "감성 분포")
    _add_bar_chart(chart_ws, 1, 17, promotion_end, "H31", "홍보성 분포")
    _add_line_chart(chart_ws, 5, 17, daily_end, "A46", "일자별 기사 추이")
    _add_bar_chart(chart_ws, 1, 32, social_end, "H46", "SNS 플랫폼별 언급")
    for col in range(1, 12):
        chart_ws.column_dimensions[chart_ws.cell(1, col).column_letter].width = 16

    priority_ws = wb.create_sheet("우선 확인 기사")
    priority_headers = ["순위", "제목", "출처", "게시 일시", "감성", "홍보성", "중요도", "AI 요약", "원문 URL"]
    priority_ws.append(priority_headers)
    for idx, row in enumerate(context["top_articles"], start=1):
        priority_ws.append([
            idx,
            row["title"] or "",
            row["publisher"] or "",
            _fmt_dt(row["published_at"]),
            _sentiment_key(row["sentiment"]),
            _promotion_label(row["is_promotion"]),
            _score_value(row) or "",
            row["summary_text"] or "",
            row["url"] or "",
        ])
    for idx, width in enumerate([8, 56, 18, 18, 12, 12, 10, 66, 52], start=1):
        priority_ws.column_dimensions[priority_ws.cell(1, idx).column_letter].width = width

    article_ws = wb.create_sheet("기사 상세")
    article_ws.append(ARTICLE_HEADERS)
    for row in rows:
        article_ws.append([
            row["id"],
            row["title"] or "",
            row["publisher"] or "",
            _fmt_dt(row["published_at"]),
            row["keywords"] or "",
            _sentiment_key(row["sentiment"]),
            _promotion_label(row["is_promotion"]),
            _score_value(row) or "",
            row["importance_reason"] or "",
            row["summary_text"] or "",
            row["url"] or "",
            row["language"] or "",
            _fmt_dt(row["created_at"]),
        ])
    for idx, width in enumerate(ARTICLE_WIDTHS, start=1):
        article_ws.column_dimensions[article_ws.cell(1, idx).column_letter].width = width
    article_ws.freeze_panes = "A2"
    article_ws.auto_filter.ref = article_ws.dimensions
    if article_ws.max_row >= 2:
        article_ws.conditional_formatting.add(
            f"H2:H{article_ws.max_row}",
            ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF", end_type="num", end_value=100, end_color="93C5FD"),
        )

    sns_ws = wb.create_sheet("SNS 신호")
    sns_ws.append(["키워드", "플랫폼", "언급 수", "긍정 힌트", "부정 힌트", "수집 시각"])
    for row in social_rows or []:
        sns_ws.append([
            row.get("keyword_text") or "",
            row.get("source") or "",
            row.get("mention_count") or 0,
            row.get("positive_hint_count") or 0,
            row.get("negative_hint_count") or 0,
            _fmt_dt(row.get("sampled_at")),
        ])
    for idx, width in enumerate([24, 16, 12, 12, 12, 20], start=1):
        sns_ws.column_dimensions[sns_ws.cell(1, idx).column_letter].width = width
    sns_ws.auto_filter.ref = sns_ws.dimensions

    for ws in wb.worksheets:
        header_row = 1
        for cell in ws[header_row]:
            if cell.value:
                cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _style_sheet(ws)

    return wb


def _html_bar_chart(title: str, rows: list[tuple[str, int]], color: str = "#60a5fa") -> str:
    data = rows[:8] or [("?곗씠???놁쓬", 0)]
    max_value = max([value for _, value in data] + [1])
    bar_rows = []
    for label, value in data:
        width = max(2, int((value / max_value) * 100)) if max_value else 2
        bar_rows.append(
            f"""
            <tr>
              <td style="width:150px;padding:7px 8px;color:#334155;font-size:12px;">{escape(str(label)[:24])}</td>
              <td style="padding:7px 8px;">
                <div style="background:#eff6ff;border-radius:999px;height:16px;overflow:hidden;">
                  <div style="background:{color};width:{width}%;height:16px;border-radius:999px;"></div>
                </div>
              </td>
              <td style="width:52px;padding:7px 8px;text-align:right;font-weight:700;color:#172033;">{value}</td>
            </tr>
"""
        )
    return f"""
    <div style="border:1px solid #dbeafe;border-radius:8px;padding:14px;margin-bottom:12px;">
      <h3 style="margin:0 0 8px;font-size:16px;color:#1e3a8a;">{escape(title)}</h3>
      <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">{''.join(bar_rows)}</table>
    </div>
    """


def _build_email_html(context: dict) -> str:
    keyword_label = escape(context["keyword_name"] or "전체 키워드")
    executive_items = "".join(f"<li>{escape(line)}</li>" for line in context["executive_summary"])
    top_rows = "".join(
        f"""
        <tr>
          <td style="padding:10px;border-bottom:1px solid #dbeafe;">{idx}</td>
          <td style="padding:10px;border-bottom:1px solid #dbeafe;"><strong>{escape(row['title'] or '제목 없음')}</strong><br>
            <span style="color:#64748b;">{escape(row['publisher'] or '출처 없음')} · {_fmt_dt(row['published_at'])}</span>
            <p style="margin:8px 0 0;color:#334155;line-height:1.5;">{escape((row['summary_text'] or '요약문이 아직 없습니다.')[:420])}</p>
          </td>
          <td style="padding:10px;border-bottom:1px solid #dbeafe;text-align:right;font-weight:700;">{(_score_value(row) or 0):.1f}</td>
        </tr>
        """
        for idx, row in enumerate(context["top_articles"][:5], start=1)
    ) or '<tr><td colspan="3" style="padding:14px;color:#64748b;">중요도 산정 완료 기사가 아직 없습니다.</td></tr>'

    return f"""
    <html>
      <body style="margin:0;background:#f8fbff;font-family:Arial,'Apple SD Gothic Neo','Malgun Gothic',sans-serif;color:#172033;">
        <div style="max-width:860px;margin:0 auto;padding:28px;">
          <div style="background:linear-gradient(135deg,#eff6ff,#ffffff);border:1px solid #dbeafe;border-radius:10px;padding:24px;">
            <div style="font-size:13px;color:#64748b;">{escape(context['generated_at'])}</div>
            <h1 style="margin:8px 0 0;font-size:26px;color:#1e3a8a;">AI 뉴스 인텔리전스 리포트</h1>
            <p style="margin:8px 0 0;color:#334155;">{keyword_label} 기준 기사, AI 분석, SNS 신호를 정리했습니다.</p>
          </div>

          <table width="100%" cellspacing="0" cellpadding="0" style="margin:18px 0;background:white;border:1px solid #dbeafe;border-radius:10px;overflow:hidden;">
            <tr>
              <td style="padding:18px;"><div style="color:#64748b;font-size:12px;">전체 기사</div><strong style="font-size:24px;">{context['total_count']}</strong></td>
              <td style="padding:18px;"><div style="color:#64748b;font-size:12px;">부정 비중</div><strong style="font-size:24px;">{context['negative_rate']}%</strong></td>
              <td style="padding:18px;"><div style="color:#64748b;font-size:12px;">홍보성 비중</div><strong style="font-size:24px;">{context['promotion_rate']}%</strong></td>
              <td style="padding:18px;"><div style="color:#64748b;font-size:12px;">SNS 언급</div><strong style="font-size:24px;">{context['social_total']}</strong></td>
            </tr>
          </table>

          <div style="background:white;border:1px solid #dbeafe;border-radius:10px;padding:16px;margin-bottom:18px;">
            <h2 style="margin:0 0 12px;font-size:18px;">Executive Summary</h2>
            <ul style="margin:0;padding-left:20px;color:#334155;line-height:1.7;">{executive_items}</ul>
          </div>

          <div style="background:white;border:1px solid #dbeafe;border-radius:10px;padding:16px;margin-bottom:18px;">
            <h2 style="margin:0 0 12px;font-size:18px;">그래프 요약</h2>
            {_html_bar_chart("감성 분포", context["sentiment_counts"], "#60a5fa")}
            {_html_bar_chart("홍보성 분포", context["promotion_counts"], "#f59e0b")}
            {_html_bar_chart("SNS 플랫폼별 언급", context["social_source_counts"], "#22d3ee")}
          </div>

          <div style="background:white;border:1px solid #dbeafe;border-radius:10px;padding:16px;">
            <h2 style="margin:0 0 12px;font-size:18px;">우선 확인 기사</h2>
            <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
              <thead><tr style="background:#eff6ff;"><th align="left" style="padding:10px;">#</th><th align="left" style="padding:10px;">기사 및 요약</th><th align="right" style="padding:10px;">중요도</th></tr></thead>
              <tbody>{top_rows}</tbody>
            </table>
          </div>
        </div>
      </body>
    </html>
    """


def _build_email_text(context: dict) -> str:
    top_lines = [
        f"{idx}. {row['title'] or '제목 없음'} / 중요도 {(_score_value(row) or 0):.1f}"
        for idx, row in enumerate(context["top_articles"][:5], start=1)
    ]
    return "\n".join([
        "AI 뉴스 인텔리전스 리포트",
        f"생성 시각: {context['generated_at']}",
        f"키워드: {context['keyword_name'] or '전체 키워드'}",
        "",
        "Executive Summary",
        *[f"- {line}" for line in context["executive_summary"]],
        "",
        f"전체 기사: {context['total_count']}",
        f"부정 비중: {context['negative_rate']}%",
        f"홍보성 비중: {context['promotion_rate']}%",
        f"SNS 언급: {context['social_total']}",
        "",
        "우선 확인 기사",
        *(top_lines or ["중요도 산정 완료 기사가 아직 없습니다."]),
        "",
        "상세 내용은 첨부된 Excel 보고서에서 확인해 주세요.",
    ])


async def _report_rows(db: AsyncSession, current_user: User, keyword_id: int | None):
    stmt = _build_query(current_user, keyword_id=keyword_id)
    result = await db.execute(stmt)
    return result.mappings().all()


def _dongguk_grouped_articles(articles: list[DonggukMailArticle]) -> dict[str, list[DonggukMailArticle]]:
    grouped: dict[str, list[DonggukMailArticle]] = {}
    for article in articles:
        grouped.setdefault(article.category or "기타", []).append(article)
    return grouped


def _dongguk_article_links(article: DonggukMailArticle) -> list[str]:
    if article.url:
        links = [article.url, *(article.links or [])]
    else:
        links = [*(article.links or [])]
    cleaned = []
    for link in links:
        value = str(link or "").strip()
        if value and value not in cleaned:
            cleaned.append(value)
    return cleaned


def _exact_dongguk_article_key(article: DonggukMailArticle) -> str:
    links = _dongguk_article_links(article)
    if links:
        return f"url:{canonicalize_article_url(links[0])}"
    title = normalize_article_title(article.title, article.source)
    source = re.sub(r"\s+", " ", (article.source or "")).strip().lower()
    if title and source:
        return f"title-source:{title}|{source}"
    if article.id is not None:
        return f"id:{article.id}"
    return f"title:{(article.title or '').strip().lower()}|{(article.source or '').strip().lower()}"


def _dedupe_exact_dongguk_articles(
    articles: list[DonggukMailArticle],
) -> tuple[list[DonggukMailArticle], list[DonggukMailArticle]]:
    kept: list[DonggukMailArticle] = []
    removed: list[DonggukMailArticle] = []
    seen: set[str] = set()
    for article in articles:
        key = _exact_dongguk_article_key(article)
        duplicate_index = next((index for index, existing in enumerate(kept) if (
            key == _exact_dongguk_article_key(existing)
            or is_same_publisher_article(
                left_title=article.title,
                left_publisher=article.source,
                left_content=article.summary,
                left_url=article.url or (article.links[0] if article.links else None),
                right_title=existing.title,
                right_publisher=existing.source,
                right_content=existing.summary,
                right_url=existing.url or (existing.links[0] if existing.links else None),
            )
        )), None)
        if key not in seen and duplicate_index is None:
            seen.add(key)
            kept.append(article)
            continue
        if duplicate_index is not None:
            representative = kept[duplicate_index]
            representative.links = list(dict.fromkeys([
                *_dongguk_article_links(representative),
                *_dongguk_article_links(article),
            ]))
            representative.is_syndicated = representative.is_syndicated or len(representative.links) > 1
        duplicate_data = article.model_dump()
        duplicate_data["selection_reason"] = "같은 언론사의 동일 원문 기사로 확인되어 대표 기사 1건에 통합했습니다."
        removed.append(DonggukMailArticle(**duplicate_data))
    return kept, removed


def _dongguk_email_html(subject: str, articles: list[DonggukMailArticle]) -> str:
    articles, _ = _dedupe_exact_dongguk_articles(articles)
    grouped = _dongguk_grouped_articles(articles)
    sections = []
    for category, rows in grouped.items():
        article_rows = []
        for idx, article in enumerate(rows, start=1):
            source = article.source or "언론사 없음"
            suffix = " 외" if article.is_syndicated else ""
            link_rows = "".join(
                f'<div style="margin-top:5px;"><a href="{escape(link)}" style="color:#1d4ed8;text-decoration:none;">{escape(link)}</a></div>'
                for link in _dongguk_article_links(article)
            )
            article_rows.append(
                f"""
                <div style="border-top:1px solid #e5e7eb;padding:15px 0;">
                  <div style="font-size:16px;font-weight:700;line-height:1.5;color:#111827;">
                    {idx}. {escape(article.title)} [{escape(source)}]{suffix}
                  </div>
                  <p style="margin:9px 0 0;color:#374151;line-height:1.65;">{escape(article.summary or '요약문이 아직 없습니다.')}</p>
                  <div style="margin-top:9px;color:#4b5563;font-size:13px;">{link_rows}</div>
                </div>
                """
            )
        sections.append(
            f"""
            <section style="background:#ffffff;border:1px solid #d1d5db;border-radius:6px;margin-top:16px;padding:16px;">
              <h2 style="font-size:18px;line-height:1.4;margin:0 0 4px;color:#1f2937;">{escape(category)}</h2>
              <div style="color:#6b7280;font-size:13px;margin-bottom:6px;">{len(rows)}건</div>
              {''.join(article_rows)}
            </section>
            """
        )
    return f"""
    <html>
      <body style="background:#f3f4f6;margin:0;padding:24px;font-family:Arial,'Malgun Gothic','Apple SD Gothic Neo',sans-serif;color:#111827;">
        <main style="max-width:820px;margin:0 auto;">
          <header style="background:#ffffff;border:1px solid #d1d5db;border-radius:6px;padding:20px;">
            <h1 style="font-size:24px;line-height:1.35;margin:6px 0 0;">{escape(subject)}</h1>
            <p style="margin:12px 0 0;color:#4b5563;line-height:1.55;">
              관리자 확인 대시보드:
              <a href="{escape(DONGGUK_DASHBOARD_URL)}" style="color:#1d4ed8;text-decoration:none;">{escape(DONGGUK_DASHBOARD_URL)}</a>
            </p>
          </header>
          {''.join(sections)}
        </main>
      </body>
    </html>
    """


def _dongguk_email_text(subject: str, articles: list[DonggukMailArticle]) -> str:
    lines = [subject, f"대시보드: {DONGGUK_DASHBOARD_URL}", ""]
    for category, rows in _dongguk_grouped_articles(articles).items():
        lines.extend([f"[{category}]", ""])
        for idx, article in enumerate(rows, start=1):
            suffix = " 외" if article.is_syndicated else ""
            lines.append(f"{idx}. {article.title} [{article.source or '언론사 없음'}]{suffix}")
            if article.summary:
                lines.append(article.summary)
            lines.extend(_dongguk_article_links(article))
            lines.append("")
    return "\n".join(lines).strip()


def _dongguk_section_label(section: str | None) -> str:
    value = str(section or "")
    if value in {"dongguk_core", "foundation"}:
        return "동국대 [법인/건학위] 관련 기사"
    if "교육" in value:
        return "대학 [교육] 관련 기사"
    if value == "education":
        return "대학 [교육] 관련 기사"
    if "종단" in value or "불교" in value or value == "buddhism":
        return "불교 [종단] 관련 기사"
    return "동국대 [법인/건학위] 관련 기사"


def _dongguk_article_response(article: DonggukMailArticle) -> dict:
    data = article.model_dump()
    data["section_label"] = _dongguk_section_label(article.section)
    return data


def _dongguk_section_key(section: str | None) -> str:
    label = _dongguk_section_label(section)
    if "교육" in label:
        return "education"
    if "불교" in label or "종단" in label:
        return "buddhism"
    return "foundation"


def _dongguk_article_qualification_text(article: DonggukMailArticle) -> str:
    return f"{article.title or ''} {article.summary or ''} {article.source or ''} {article.category or ''}"


def _has_hongbo_hard_exclusion(article: DonggukMailArticle, section: str) -> bool:
    text = _dongguk_article_qualification_text(article)
    title = article.title or ""
    url = article.url or (article.links[0] if article.links else "")

    if section == "foundation":
        if re.search(r"대학\s*뉴스브리핑|뉴스브리핑", title, re.I):
            return True
        if re.search(r"불꽃야구|직관\s*경기|프로야구|연예\s*예능", text, re.I):
            return True
        if re.search(r"동국대.*(제압|대파|꺾고|패배|탈락)|상대.*동국대.*(승|제압)", text, re.I):
            return True
        if re.search(r"신정아|패소|소송", text, re.I):
            return True
        return False

    if section == "education":
        if re.search(r"\[[^\]]*수시|수시특집|입시특집|개별\s*대학|호서대|청주대|경복대|서울시립대|대구가톨릭대", title, re.I):
            return True
        if re.search(r"초등학교|중학교|고등학교|영유아|유치원|어린이집", text, re.I) and not re.search(
            r"고등교육|대학|교부금|교육부|대교협|사총협", text, re.I
        ):
            return True
        return False

    if section == "buddhism":
        if re.search(r"기자칼럼|맛집|여행|지역축제|기획전|합창단", title, re.I):
            return True
        if re.search(r"개별\s*사찰\s*행사", text, re.I) and not re.search(
            r"조계종|종단|포교|출가|성보|불교문화유산|불교중앙박물관", text, re.I
        ):
            return True
        return False

    return False


def _is_foundation_mail_eligible(article: DonggukMailArticle) -> bool:
    if _has_hongbo_hard_exclusion(article, "foundation"):
        return False
    text = _dongguk_article_qualification_text(article)
    title = article.title or ""
    source = article.source or ""
    has_direct_title_signal = bool(re.search(r"동국대|동국대학교|동국대WISE|WISE캠퍼스|동문소식", title, re.I))
    has_official_source_signal = source.strip() == "동국대학교"
    if not (has_direct_title_signal or has_official_source_signal):
        return False
    positive = (
        r"총장|이사장|부총장|교수|학생|동문|학부|학과|대학원|캠퍼스|기술지주|"
        r"연구소|학술원|종학연구소|불교학술원|기부|장학|발전기금|로터스관|"
        r"협약|사업|선정|수상|인증|개최|성료|모집|임용|선임|취임|연구|개발|"
        r"AI|C포럼|서울RISE|농구|MBC배|입시|교사연수|Dream"
    )
    if not re.search(positive, text, re.I):
        return False
    return True


def _is_education_mail_eligible(article: DonggukMailArticle) -> bool:
    if _has_hongbo_hard_exclusion(article, "education"):
        return False
    text = _dongguk_article_qualification_text(article)
    return bool(
        re.search(
            r"대학|고등교육|교육부|대교협|수능|대입|입시|등록금|교부금|학사|"
            r"정원|국공립대|사립대|전문대|의대|문·이과|AI\s*기본\s*역량|"
            r"교육재정|고등교육법|지역인재|연구원|교원창업",
            text,
            re.I,
        )
    )


def _is_buddhism_mail_eligible(article: DonggukMailArticle) -> bool:
    if _has_hongbo_hard_exclusion(article, "buddhism"):
        return False
    text = _dongguk_article_qualification_text(article)
    return bool(
        re.search(
            r"불교|조계종|종단|스님|사찰|성보|포교|선명상|수행|출가|법회|"
            r"불상|불두|불교문화유산|불교중앙박물관|불교신문|법보신문|현대불교|BTN|BBS",
            text,
            re.I,
        )
    )


def _is_dongguk_mail_section_eligible(article: DonggukMailArticle) -> bool:
    section = _dongguk_section_key(article.section)
    if section == "education":
        return _is_education_mail_eligible(article)
    if section == "buddhism":
        return _is_buddhism_mail_eligible(article)
    return _is_foundation_mail_eligible(article)


def _dongguk_candidate_section_key(article: DonggukMailArticle) -> str | None:
    if _is_foundation_mail_eligible(article):
        title = article.title or ""
        source = article.source or ""
        url = article.url or (article.links[0] if article.links else "")
        if re.search(r"동국대|동국대학교|동국대WISE|WISE캠퍼스|동문소식", title, re.I) or source.strip() == "동국대학교" or "dongguk.edu" in url:
            return "foundation"
    raw_section = str(article.section or "").strip()
    if raw_section:
        return _dongguk_section_key(raw_section)
    if _is_education_mail_eligible(article):
        return "education"
    if _is_buddhism_mail_eligible(article):
        return "buddhism"
    if _is_foundation_mail_eligible(article):
        return "foundation"
    return None


def _score_pattern(text: str, pattern: str, weight: int) -> int:
    return weight if re.search(pattern, text, re.I) else 0


def _source_contains(article: DonggukMailArticle, pattern: str) -> bool:
    return bool(re.search(pattern, article.source or "", re.I))


def _hongbo_recency_score(article: DonggukMailArticle, mail_date: str | None) -> int:
    if not mail_date or not article.published_at:
        return 0
    try:
        published_at = datetime.fromisoformat(str(article.published_at).replace("Z", "+00:00"))
        if published_at.tzinfo is None:
            published_at = published_at.replace(tzinfo=KST)
        published_at = published_at.astimezone(KST)
        window_start, window_end = _dongguk_report_window(mail_date, time(8, 30))
        if window_start <= published_at <= window_end:
            return 70
        mail_day = datetime.fromisoformat(mail_date).date()
        day_delta = (mail_day - published_at.date()).days
        if day_delta == 0:
            return 45
        if day_delta == 1:
            return 55
        if day_delta == 2:
            return 20
        if day_delta >= 4:
            return -85
        if day_delta < 0:
            return -20
    except Exception:
        return 0
    return 0


def _hongbo_selection_score(
    article: DonggukMailArticle,
    section: str,
    editor_selected: bool = False,
    mail_date: str | None = None,
) -> float:
    if _has_hongbo_hard_exclusion(article, section):
        return -1000
    text = _dongguk_article_qualification_text(article)
    title = article.title or ""
    url = article.url or (article.links[0] if article.links else "")
    score = float(article.score or 0) * 0.25
    score += 8 if editor_selected else 0
    score += _hongbo_recency_score(article, mail_date)

    if section == "foundation":
        score += _score_pattern(title, r"동국대|동국대학교|동국대WISE|WISE캠퍼스", 35)
        score += 30 if _source_contains(article, r"동국대학교") or "dongguk.edu" in url else 0
        score += _score_pattern(title, r"C포럼|\[2026\s*C포럼\]|총장|이사장", 80)
        score += _score_pattern(text, r"C포럼|총장|이사장|부총장|공식\s*메시지", 44)
        score += _score_pattern(text, r"기부|장학|발전기금|로터스관|희사", 32)
        score += _score_pattern(text, r"동문소식|동문|명예회장|대표이사|회장|본부장|배우|감독", 26)
        score += _score_pattern(text, r"지능\s*IoT|공동학위", 68)
        score += _score_pattern(text, r"대한민국학술원|학술원\s*신임회원|황훈성", 62)
        score += _score_pattern(text, r"서울국제명상엑스포|명상엑스포", 62)
        score += _score_pattern(text, r"중구불교협의회|하안거|선.?교\s*겸수", 50)
        score += _score_pattern(text, r"간화선|선수행|집중수행", 34)
        score += _score_pattern(text, r"기술지주|모태펀드|대학창업펀드|창업펀드|운용사|사업.*선정|과제.*선정|앵커|RISE", 24)
        score += _score_pattern(text, r"연구|개발|학술원|종학연구소|불교학술원|명상엑스포|국제학술대회|학술원\s*신임회원", 22)
        score += _score_pattern(text, r"핀테크|블록체인", 20)
        score += _score_pattern(text, r"Dream Workshop|교사연수|공식행사|성료|개최", 14)
        score += _score_pattern(text, r"MBC배|농구|결선|우승|역전승|한양대", 18)
        score -= _score_pattern(text, r"대학\s*뉴스브리핑|뉴스브리핑|수시특집|입시특집|안전보건공시제|임상통계\s*소프트웨어|베트남\s*유학생", 36)
        score -= _score_pattern(text, r"패소|신정아|불꽃야구|패배|탈락|부상|상대.*동국대.*제압|동국대에.*승|동국대.*대파", 80)
    elif section == "education":
        score += 28 if _source_contains(article, r"한국대학신문|뉴스1|파이낸셜뉴스|뉴시스|연합뉴스|중앙일보") else 0
        score += _score_pattern(text, r"고등교육재정교부금법|교육교부금|교부금.*개편|고등교육.*투자|사총협", 36)
        score += _score_pattern(text, r"교육부|대교협|고등교육법|시행령|국공립대|사립대|등록금|AI\s*기본\s*역량|문·이과|지역인재|입학\s*취소", 26)
        score += _score_pattern(text, r"교원창업", 48)
        score += _score_pattern(text, r"대입|수능|수시|개인정보.*전송|마이데이터", 12)
        score -= _score_pattern(text, r"포토|\[[^\]]*수시|수시특집|개별\s*대학|호서대|청주대|경복대|서울시립대|대구가톨릭대|동국대\s*WISE", 42)
        score -= _score_pattern(text, r"초중고|영유아|고등학교|중학교|초등", 12)
    elif section == "buddhism":
        score += 28 if _source_contains(article, r"불교신문|법보신문|BBS|불교방송|BTN|현대불교|대한불교조계종") else 0
        score += _score_pattern(text, r"조계종|종단|포교|포교사단|출가|출가상담|정광고|종책|총무원장", 34)
        score += _score_pattern(text, r"성보|불교문화유산|불교중앙박물관|탑비|벽화|불두|복장물|비파괴조사|보존처리|국보|보물", 30)
        score += _score_pattern(text, r"부산불교박람회|불교박람회|부산.*벡스코|색즉시공|공놀이|청년|AI|종교.*AI|불교.*AI|만공|포산|근일\s*대종사|법맥", 30)
        score += _score_pattern(text, r"스님|사찰|수행|선명상|법회", 8)
        score -= _score_pattern(text, r"기자칼럼|여행|맛집|영화제|인터뷰|기획전|지역축제|합창단|개별\s*사찰\s*행사", 24)

    return score


def _article_with_mail_section(article: DonggukMailArticle, section: str) -> DonggukMailArticle:
    desired = {
        "foundation": "dongguk_core",
        "education": "education",
        "buddhism": "buddhism",
    }[section]
    if article.section == desired:
        return article
    data = article.model_dump()
    data["section"] = desired
    return DonggukMailArticle(**data)


def _dongguk_mail_section_policy(
    selected_articles: list[DonggukMailArticle],
    excluded_articles: list[DonggukMailArticle],
    all_candidates: list[DonggukMailArticle],
    mail_date: str | None = None,
) -> tuple[list[DonggukMailArticle], list[DonggukMailArticle]]:
    section_targets = {
        "foundation": {"min": 4, "max": 6},
        "education": {"min": 2, "max": 2},
        "buddhism": {"min": 2, "max": 2},
    }
    selected_articles, selected_exact_duplicates = _dedupe_exact_dongguk_articles(selected_articles)
    selected_articles, selected_topic_duplicates = _dedupe_representative_articles(selected_articles)
    all_candidates, candidate_exact_duplicates = _dedupe_exact_dongguk_articles(all_candidates)
    candidate_topic_duplicates: list[DonggukMailArticle] = []
    selected_keys = {_dongguk_article_key(article) for article in selected_articles}
    excluded_by_key = {_dongguk_article_key(article): article for article in excluded_articles}

    qualified_pool: dict[str, list[DonggukMailArticle]] = {key: [] for key in section_targets}
    disqualified_selected: list[DonggukMailArticle] = []
    for article in all_candidates:
        section = _dongguk_candidate_section_key(article)
        if section not in qualified_pool:
            continue
        if (
            (section == "foundation" and _is_foundation_mail_eligible(article))
            or (section == "education" and _is_education_mail_eligible(article))
            or (section == "buddhism" and _is_buddhism_mail_eligible(article))
        ):
            qualified_pool[section].append(article)

    for section, articles in qualified_pool.items():
        articles.sort(
            key=lambda article: (
                _hongbo_selection_score(article, section, _dongguk_article_key(article) in selected_keys, mail_date),
                str(article.published_at or ""),
            ),
            reverse=True,
        )

    final_articles: list[DonggukMailArticle] = []
    used_keys: set[str] = set()
    for article in selected_articles:
        if not _is_dongguk_mail_section_eligible(article):
            data = article.model_dump()
            data["selection_reason"] = "섹션 자격 필터를 통과하지 못해 메일 포함에서 제외되었습니다."
            disqualified_selected.append(DonggukMailArticle(**data))

    for section, target in section_targets.items():
        current = 0
        target_count = target["max"] if section == "foundation" else target["min"]
        for candidate in qualified_pool[section]:
            key = _dongguk_article_key(candidate)
            if key in used_keys:
                continue
            if any(
                _dongguk_section_key(selected.section) == section
                and _is_similar_dongguk_article(candidate, selected)
                for selected in final_articles
            ):
                data = candidate.model_dump()
                data["selection_reason"] = "같은 주제의 더 높은 홍보처 적합도 기사로 묶여 제외되었습니다."
                candidate_topic_duplicates.append(DonggukMailArticle(**data))
                continue
            final_articles.append(_article_with_mail_section(candidate, section))
            used_keys.add(key)
            current += 1
            if current >= target_count:
                break

    capped_articles: list[DonggukMailArticle] = []
    section_counts = {key: 0 for key in section_targets}
    over_cap: list[DonggukMailArticle] = []
    for article in final_articles:
        section = _dongguk_section_key(article.section)
        max_count = section_targets[section]["max"]
        if section_counts[section] >= max_count:
            data = article.model_dump()
            data["selection_reason"] = "홍보처 메일 섹션별 최대 수량을 초과해 제외되었습니다."
            over_cap.append(DonggukMailArticle(**data))
            continue
        capped_articles.append(article)
        section_counts[section] += 1

    excluded = [
        *excluded_by_key.values(),
        *selected_exact_duplicates,
        *selected_topic_duplicates,
        *candidate_exact_duplicates,
        *candidate_topic_duplicates,
        *disqualified_selected,
        *over_cap,
    ]
    included_keys = {_dongguk_article_key(article) for article in capped_articles}
    for article in all_candidates:
        key = _dongguk_article_key(article)
        if key in included_keys or key in excluded_by_key:
            continue
        if _is_dongguk_mail_section_eligible(article) and key not in selected_keys:
            data = article.model_dump()
            data["selection_reason"] = "섹션별 목표 수량 안에서 더 높은 우선순위 기사에 밀려 제외되었습니다."
            excluded.append(DonggukMailArticle(**data))

    deduped_excluded, _ = _dedupe_exact_dongguk_articles(excluded)
    return capped_articles, deduped_excluded


def _dongguk_section_grouped_articles(articles: list[DonggukMailArticle]) -> dict[str, list[DonggukMailArticle]]:
    order = [
        "동국대 [법인/건학위] 관련 기사",
        "대학 [교육] 관련 기사",
        "불교 [종단] 관련 기사",
    ]
    grouped = {label: [] for label in order}
    for article in articles:
        grouped.setdefault(_dongguk_section_label(article.section), []).append(article)
    return {label: grouped.get(label, []) for label in order if grouped.get(label)}


def _dongguk_hangul_article_title(index: int, article: DonggukMailArticle) -> str:
    source = (article.source or "").strip()
    suffix = " 외" if article.is_syndicated or len(_dongguk_article_links(article)) > 1 else ""
    if source.endswith(" 외"):
        source = source[:-2].strip()
        suffix = " 외"
    source_part = f" [{source}]" if source else ""
    return f"{index}. {article.title}{source_part}{suffix}"


def _set_hwpx_paragraph_text(paragraph: ET.Element, text: str) -> ET.Element:
    text_nodes = paragraph.findall(".//hp:t", HWPX_NS)
    if not text_nodes:
        return paragraph
    text_nodes[0].text = text
    for node in text_nodes[1:]:
        node.text = ""
    return paragraph


def _dongguk_hwpx_preview_text(subject: str, articles: list[DonggukMailArticle]) -> str:
    lines = [f"<{subject}><><>", "<><>", "대외협력처 홍보실", ""]
    for section, rows in _dongguk_section_grouped_articles(articles).items():
        lines.append(section)
        for idx, article in enumerate(rows, start=1):
            lines.append(_dongguk_hangul_article_title(idx, article))
            for link in _dongguk_article_links(article):
                lines.append(f"{link} ")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def _dongguk_hwpx_bytes(subject: str, articles: list[DonggukMailArticle]) -> bytes:
    articles, _ = _dedupe_exact_dongguk_articles(articles)
    if not DONGGUK_HWPX_TEMPLATE.exists():
        raise FileNotFoundError("Dongguk HWPX template is missing.")

    ET.register_namespace("hs", HWPX_NS["hs"])
    ET.register_namespace("hp", HWPX_NS["hp"])

    with zipfile.ZipFile(DONGGUK_HWPX_TEMPLATE, "r") as template:
        section_xml = template.read("Contents/section0.xml")
        root = ET.fromstring(section_xml)
        original_children = list(root)
        title_proto = original_children[0]
        office_proto = original_children[1]
        spacer_proto = original_children[2]
        section_proto = original_children[3]
        line_proto = original_children[4]
        blank_proto = original_children[14] if len(original_children) > 14 else original_children[2]

        for child in list(root):
            root.remove(child)

        root.append(_set_hwpx_paragraph_text(deepcopy(title_proto), subject))
        root.append(_set_hwpx_paragraph_text(deepcopy(office_proto), "대외협력처 홍보실"))
        root.append(_set_hwpx_paragraph_text(deepcopy(spacer_proto), ""))

        for section, rows in _dongguk_section_grouped_articles(articles).items():
            root.append(_set_hwpx_paragraph_text(deepcopy(section_proto), section))
            for idx, article in enumerate(rows, start=1):
                root.append(_set_hwpx_paragraph_text(deepcopy(line_proto), _dongguk_hangul_article_title(idx, article)))
                for link in _dongguk_article_links(article):
                    root.append(_set_hwpx_paragraph_text(deepcopy(line_proto), f"{link} "))
            root.append(_set_hwpx_paragraph_text(deepcopy(blank_proto), ""))

        new_section = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        preview_text = _dongguk_hwpx_preview_text(subject, articles).encode("utf-8")
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w") as generated:
            for info in template.infolist():
                if info.filename in {"Contents/section0.xml", "Preview/PrvText.txt"}:
                    continue
                data = template.read(info.filename)
                compression = zipfile.ZIP_STORED if info.filename == "mimetype" else zipfile.ZIP_DEFLATED
                generated.writestr(info.filename, data, compress_type=compression)
            generated.writestr("Contents/section0.xml", new_section, compress_type=zipfile.ZIP_DEFLATED)
            generated.writestr("Preview/PrvText.txt", preview_text, compress_type=zipfile.ZIP_DEFLATED)
        return output.getvalue()


def _dongguk_hwpx_filename(subject: str, mail_date: str | None = None) -> str:
    filename_date = mail_date or _mail_date_from_subject(subject) or datetime.now(KST).date().isoformat()
    compact_date = filename_date.replace("-", "")[2:] if re.match(r"^\d{4}-\d{2}-\d{2}$", filename_date) else filename_date
    return f"오늘의 주요 뉴스({compact_date}).hwpx"


def _normalize_title(value: str | None) -> str:
    return "".join(ch for ch in (value or "").lower() if ch.isalnum())


def _is_similar_title(left: str, right: str) -> bool:
    left_key = _normalize_title(left)
    right_key = _normalize_title(right)
    if not left_key or not right_key:
        return False
    if left_key in right_key or right_key in left_key:
        return True
    return SequenceMatcher(None, left_key, right_key).ratio() >= 0.72


def _mail_date_from_subject(subject: str | None) -> str | None:
    match = re.search(r"(\d{4})\.(\d{2})\.(\d{2})", subject or "")
    if not match:
        return None
    return "-".join(match.groups())


async def _sent_dongguk_titles(db: AsyncSession, user_id: int, before_mail_date: str | None = None) -> set[str]:
    stmt = (
        select(EmailDelivery.body)
        .where(EmailDelivery.user_id == user_id)
        .where(EmailDelivery.status == "SENT")
        .order_by(EmailDelivery.sent_at.desc().nullslast(), EmailDelivery.created_at.desc())
        .limit(120)
    )
    result = await db.execute(stmt)
    titles: set[str] = set()
    for body in result.scalars().all():
        try:
            payload = json.loads(body or "{}")
        except json.JSONDecodeError:
            continue
        if payload.get("type") != "DONGGUK_PR_DAILY":
            continue
        sent_mail_date = payload.get("mail_date") or _mail_date_from_subject(payload.get("subject"))
        if before_mail_date and sent_mail_date and sent_mail_date >= before_mail_date:
            continue
        for article in payload.get("articles") or []:
            title = article.get("title")
            if title:
                titles.add(title)
    return titles


def _filter_previously_sent_articles(
    articles: list[DonggukMailArticle],
    sent_titles: set[str],
) -> tuple[list[DonggukMailArticle], list[DonggukMailArticle]]:
    included: list[DonggukMailArticle] = []
    excluded: list[DonggukMailArticle] = []
    for article in articles:
        if any(_is_similar_title(article.title, sent_title) for sent_title in sent_titles):
            excluded.append(article)
        else:
            included.append(article)
    return included, excluded


def _dongguk_delivery_body(
    subject: str,
    articles: list[DonggukMailArticle],
    excluded: list[DonggukMailArticle],
    mail_date: str | None = None,
    keyword_id: int | None = None,
) -> str:
    articles, _ = _dedupe_exact_dongguk_articles(articles)
    return json.dumps(
        {
            "type": "DONGGUK_PR_DAILY",
            "subject": subject,
            "mail_date": mail_date or _mail_date_from_subject(subject),
            "keyword_id": keyword_id,
            "articles": [article.model_dump() for article in articles],
            "excluded_articles": [article.model_dump() for article in excluded],
        },
        ensure_ascii=False,
    )


def _dongguk_mail_subject(mail_date: str) -> str:
    parsed = datetime.strptime(mail_date, "%Y-%m-%d").date()
    weekdays = ["월", "화", "수", "목", "금", "토", "일"]
    return f"오늘의 주요 뉴스 {parsed:%Y.%m.%d}.[{weekdays[parsed.weekday()]}]"


def _parse_report_send_time(value: str | None) -> time:
    try:
        hour_text, minute_text = (value or "08:30")[:5].split(":")
        return time(hour=int(hour_text), minute=int(minute_text))
    except Exception:
        return time(hour=8, minute=30)


def _dongguk_report_window(mail_date: str, send_time: str | None = None) -> tuple[datetime, datetime]:
    target_date = datetime.strptime(mail_date, "%Y-%m-%d").date()
    end_at = datetime.combine(target_date, _parse_report_send_time(send_time), tzinfo=KST)
    return end_at - timedelta(days=1), end_at


def _article_item_to_dongguk(item: dict) -> DonggukMailArticle:
    url = item.get("original_url") or item.get("url") or ""
    return DonggukMailArticle(
        id=item.get("id"),
        title=item.get("title") or "제목 없음",
        source=item.get("source") or item.get("publisher") or "언론사 없음",
        collection_source=item.get("collection_source"),
        section=item.get("section"),
        category=item.get("category") or "기타",
        summary=item.get("summary") or "요약문이 아직 없습니다.",
        url=url,
        thumbnail_url=item.get("thumbnail_url"),
        published_at=str(item.get("published_at") or ""),
        links=[url] if url else [],
        score=item.get("importance") or item.get("priority_boost"),
    )


def _extract_thumbnail_url(item: dict) -> str | None:
    candidates = [
        item.get("thumbnail_url"),
        item.get("image_url"),
        item.get("image"),
        item.get("thumbnail"),
        item.get("og_image"),
        item.get("lead_image"),
        item.get("main_image"),
    ]
    images = item.get("images")
    if isinstance(images, list):
        candidates.extend(images)
    elif isinstance(images, dict):
        candidates.extend(images.values())
    for value in candidates:
        if isinstance(value, dict):
            value = value.get("url") or value.get("src")
        if not value:
            continue
        url = str(value).strip()
        if url.startswith("http://") or url.startswith("https://"):
            return url
    return None


async def _dongguk_articles_for_keyword_date(
    db: AsyncSession,
    *,
    user_id: int,
    keyword_id: int,
    mail_date: str,
) -> list[DonggukMailArticle]:
    try:
        target_date = datetime.strptime(mail_date, "%Y-%m-%d").date()
    except ValueError:
        target_date = datetime.now(KST).date()
    work_window = await HolidayService(db).work_window(user_id, target_date)
    keyword = await db.get(Keyword, keyword_id)
    send_time = keyword.email_send_time if keyword else "08:30"
    published_from, published_to = article_collection_window(
        work_window["start_date"],
        work_window["end_date"],
        send_time,
    )
    service = ArticleService(db)
    page = 1
    items = []
    while True:
        query = app.schemas.articles.ArticleListQuery(
            page=page,
            size=100,
            keyword_id=keyword_id,
            sort=app.schemas.articles.ArticleSort.importance_desc,
            from_at=published_from,
            to_at=published_to,
        )
        page_items, total = await service.get_article_list(user_id=user_id, query=query)
        items.extend(page_items)
        if len(items) >= total or not page_items:
            break
        page += 1
    return [_article_item_to_dongguk(item) for item in items]


def _articles_for_news_editor(articles: list[DonggukMailArticle]) -> str:
    payload = []
    for index, article in enumerate(articles, start=1):
        article_id = article.id or index
        links = _dongguk_article_links(article)
        payload.append(
            {
                "id": article_id,
                "title": article.title,
                "source": article.source,
                "summary": article.summary,
                "url": article.url or (article.links[0] if article.links else ""),
                "canonical_url": canonicalize_article_url(article.url or (article.links[0] if article.links else "")),
                "original_urls": links,
                "normalized_title": normalize_article_title(article.title, article.source),
                "content_excerpt": (article.summary or "")[:4000],
                "thumbnail_url": article.thumbnail_url,
                "published_at": article.published_at,
                "section": article.section,
                "section_label": _dongguk_section_label(article.section),
                "category": article.category,
            }
        )
    return json.dumps(payload, ensure_ascii=False)


def _dongguk_article_key(article: DonggukMailArticle, index: int = 0) -> str:
    return str(article.id or article.url or article.title or index)


def _dongguk_preview_cache_key(
    *,
    user_id: int,
    keyword_id: int | None,
    mail_date: str | None,
    subject: str,
    articles: list[DonggukMailArticle],
    exclude_similar_sent: bool,
    priority_criteria: str | None = None,
) -> str:
    articles, _ = _dedupe_exact_dongguk_articles(articles)
    article_payload = []
    for article in articles:
        article_payload.append(
            {
                "id": article.id,
                "title": article.title,
                "url": article.url,
                "links": sorted(_dongguk_article_links(article)),
                "category": article.category,
                "section": article.section,
            }
        )
    payload = {
        "version": 2,
        "user_id": user_id,
        "keyword_id": keyword_id,
        "mail_date": mail_date,
        "subject": subject,
        "exclude_similar_sent": exclude_similar_sent,
        "priority_criteria": _normalize_dongguk_priority_criteria(priority_criteria),
        "articles": sorted(article_payload, key=lambda item: (str(item.get("id") or ""), item.get("title") or "")),
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


async def _get_cached_dongguk_preview(db: AsyncSession, cache_key: str) -> dict | None:
    result = await db.execute(
        select(DonggukPreviewCache.response_body)
        .where(DonggukPreviewCache.cache_key == cache_key)
        .limit(1)
    )
    body = result.scalar_one_or_none()
    if not body:
        return None
    try:
        cached = json.loads(body)
    except json.JSONDecodeError:
        return None
    if cached.get("policy_version") != DONGGUK_MAIL_POLICY_VERSION:
        return None
    if cached.get("editor_used") is False:
        return None
    cached["cached"] = True
    return cached


async def _save_dongguk_preview_cache(
    db: AsyncSession,
    *,
    user_id: int,
    keyword_id: int | None,
    mail_date: str | None,
    subject: str,
    cache_key: str,
    data: dict,
) -> None:
    stmt = pg_insert(DonggukPreviewCache).values(
        user_id=user_id,
        keyword_id=keyword_id,
        mail_date=mail_date,
        subject=subject,
        cache_key=cache_key,
        response_body=json.dumps(data, ensure_ascii=False),
    ).on_conflict_do_update(
        index_elements=[DonggukPreviewCache.cache_key],
        set_={
            "keyword_id": keyword_id,
            "mail_date": mail_date,
            "subject": subject,
            "response_body": json.dumps(data, ensure_ascii=False),
            "updated_at": datetime.now(timezone.utc),
        },
    )
    await db.execute(stmt)
    await db.commit()


def _draft_response(draft: DonggukMailDraft | None) -> dict:
    if draft is None:
        return {"found": False}
    try:
        selected_article_keys = json.loads(draft.selected_article_keys or "[]")
    except json.JSONDecodeError:
        selected_article_keys = []
    try:
        selected_articles = json.loads(draft.selected_articles or "[]")
    except json.JSONDecodeError:
        selected_articles = []
    try:
        removed_article_keys = json.loads(draft.removed_article_keys or "[]")
    except json.JSONDecodeError:
        removed_article_keys = []
    try:
        removed_articles = json.loads(draft.removed_articles or "[]")
    except json.JSONDecodeError:
        removed_articles = []
    try:
        preview_data = json.loads(draft.preview_body) if draft.preview_body else None
    except json.JSONDecodeError:
        preview_data = None
    if preview_data and preview_data.get("policy_version") != DONGGUK_MAIL_POLICY_VERSION:
        preview_data = None
    # Old drafts can contain Google/Naver URL variants of the same publisher article.
    # Sanitize them on every read so every Hongbo screen uses the same representative list.
    selected_models = [DonggukMailArticle(**item) for item in selected_articles]
    selected_models, newly_removed = _dedupe_exact_dongguk_articles(selected_models)
    selected_articles = [_dongguk_article_response(item) for item in selected_models]
    if newly_removed:
        removed_models = [DonggukMailArticle(**item) for item in removed_articles]
        removed_models.extend(newly_removed)
        removed_articles = [_dongguk_article_response(item) for item in removed_models]
    selected_article_keys = [_dongguk_article_key(item, index) for index, item in enumerate(selected_models)]
    if preview_data:
        preview_data["articles"] = selected_articles
        preview_data["excluded_articles"] = removed_articles
        preview_data["article_count"] = len(selected_articles)
        preview_data["excluded_count"] = len(removed_articles)
    return {
        "found": True,
        "id": draft.id,
        "keyword_id": draft.keyword_id,
        "mail_date": draft.mail_date,
        "subject": draft.subject,
        "selected_article_keys": selected_article_keys,
        "selected_articles": selected_articles,
        "removed_article_keys": removed_article_keys,
        "removed_articles": removed_articles,
        "preview_data": preview_data,
        "updated_at": _fmt_dt(draft.updated_at),
    }


async def _get_dongguk_mail_draft(
    db: AsyncSession,
    *,
    user_id: int,
    keyword_id: int | None,
    mail_date: str,
) -> DonggukMailDraft | None:
    result = await db.execute(
        select(DonggukMailDraft)
        .where(
            DonggukMailDraft.user_id == user_id,
            DonggukMailDraft.keyword_id == keyword_id,
            DonggukMailDraft.mail_date == mail_date,
        )
        .limit(1)
    )
    return result.scalar_one_or_none()


async def _save_dongguk_mail_draft(
    db: AsyncSession,
    *,
    user_id: int,
    keyword_id: int | None,
    mail_date: str,
    subject: str,
    selected_article_keys: list[str],
    selected_articles: list[DonggukMailArticle],
    removed_article_keys: list[str] | None = None,
    removed_articles: list[DonggukMailArticle] | None = None,
    preview_data: dict | None = None,
    record_feedback: bool = False,
    feedback_source: str = "unknown",
) -> DonggukMailDraft:
    draft = await _get_dongguk_mail_draft(db, user_id=user_id, keyword_id=keyword_id, mail_date=mail_date)
    before_selected_articles: list[dict] = []
    before_preview_data: dict | None = None
    if draft is not None and record_feedback:
        try:
            before_selected_articles = json.loads(draft.selected_articles or "[]")
        except json.JSONDecodeError:
            before_selected_articles = []
        try:
            before_preview_data = json.loads(draft.preview_body) if draft.preview_body else None
        except json.JSONDecodeError:
            before_preview_data = None
    if draft is None:
        draft = DonggukMailDraft(
            user_id=user_id,
            keyword_id=keyword_id,
            mail_date=mail_date,
            subject=subject,
        )
        db.add(draft)
    draft.subject = subject
    draft.selected_article_keys = json.dumps(selected_article_keys, ensure_ascii=False)
    draft.selected_articles = json.dumps([article.model_dump() for article in selected_articles], ensure_ascii=False)
    draft.removed_article_keys = json.dumps(removed_article_keys or [], ensure_ascii=False)
    draft.removed_articles = json.dumps([article.model_dump() for article in (removed_articles or [])], ensure_ascii=False)
    draft.preview_body = json.dumps(preview_data, ensure_ascii=False) if preview_data is not None else None
    if record_feedback and before_selected_articles:
        await PriorityInsightService(db).record_draft_changes(
            user_id=user_id,
            keyword_id=keyword_id,
            mail_date=mail_date,
            source_screen=feedback_source,
            before_selected_articles=before_selected_articles,
            before_preview_data=before_preview_data,
            after_selected_articles=[article.model_dump() for article in selected_articles],
            after_preview_data=preview_data,
        )
    await db.commit()
    await db.refresh(draft)
    return draft


def _articles_from_editor_result(result: dict, fallback_articles: list[DonggukMailArticle]) -> tuple[list[DonggukMailArticle], list[DonggukMailArticle]]:
    fallback_by_key: dict[str, DonggukMailArticle] = {}
    for index, article in enumerate(fallback_articles, start=1):
        keys = {str(article.id or index), str(index), article.title or "", article.url or ""}
        keys.update(article.links or [])
        for key in keys:
            if key:
                fallback_by_key[key] = article

    def find_fallback(*values: object) -> DonggukMailArticle | None:
        for value in values:
            if value is None:
                continue
            match = fallback_by_key.get(str(value))
            if match:
                return match
        return None

    selected: list[DonggukMailArticle] = []
    for item in result.get("selected_articles") or []:
        representative_id = item.get("representative_article_id")
        links = item.get("related_links") or []
        main_url = item.get("main_url")
        fallback = find_fallback(representative_id, main_url, item.get("title"), *(links or []))
        if main_url and main_url not in links:
            links = [main_url, *links]
        if not links and fallback:
            links = fallback.links
        selected.append(
            DonggukMailArticle(
                id=representative_id if isinstance(representative_id, int) else (fallback.id if fallback else None),
                title=item.get("title") or (fallback.title if fallback else "제목 없음"),
                source=item.get("source") or (fallback.source if fallback else None),
                section=fallback.section if fallback else None,
                category=item.get("category") or (fallback.category if fallback else "기타"),
                summary=item.get("summary") or (fallback.summary if fallback else None),
                url=main_url or (fallback.url if fallback else None),
                thumbnail_url=fallback.thumbnail_url if fallback else None,
                published_at=fallback.published_at if fallback else None,
                links=list(dict.fromkeys([link for link in links if link])),
                priority=item.get("priority") or (fallback.priority if fallback else None),
                is_syndicated=len(links) > 1,
                selection_reason=item.get("selection_reason") or item.get("priority_reason") or (fallback.selection_reason if fallback else None),
            )
        )

    excluded: list[DonggukMailArticle] = []
    for item in result.get("excluded_articles") or []:
        article_id = item.get("article_id") or item.get("representative_article_id")
        fallback = find_fallback(article_id, item.get("main_url"), item.get("title"))
        excluded.append(
            DonggukMailArticle(
                id=article_id if isinstance(article_id, int) else (fallback.id if fallback else None),
                title=item.get("title") or (fallback.title if fallback else "제외 기사"),
                source=fallback.source if fallback else None,
                section=fallback.section if fallback else None,
                category=fallback.category if fallback else "기타",
                summary=item.get("reason") or (fallback.summary if fallback else None),
                thumbnail_url=fallback.thumbnail_url if fallback else None,
                links=fallback.links if fallback else [],
                priority=fallback.priority if fallback else None,
                priority_name=fallback.priority_name if fallback else None,
                score=fallback.score if fallback else None,
                is_syndicated=fallback.is_syndicated if fallback else False,
                selection_reason=item.get("reason") or item.get("selection_reason") or (fallback.selection_reason if fallback else None),
            )
        )

    return selected or fallback_articles, excluded


async def _run_news_editor_or_fallback(
    *,
    current_user: User,
    mail_date: str,
    subject: str,
    articles: list[DonggukMailArticle],
    priority_criteria: str | None = None,
) -> tuple[list[DonggukMailArticle], list[DonggukMailArticle], dict | None]:
    try:
        result = await DifyService.from_settings().run_news_editor_workflow(
            mail_date=mail_date,
            subject=subject,
            articles_json=_articles_for_news_editor(articles),
            priority_criteria=_normalize_dongguk_priority_criteria(priority_criteria),
            user=f"user-{current_user.id}",
        )
        selected, excluded = _articles_from_editor_result(result, articles)
        return selected, excluded, result
    except Exception as exc:
        print(f"Dongguk news editor workflow failed: {exc}")
        raise HTTPException(
            status_code=503,
            detail="AI 기사 선정에 실패했습니다. AI 연결 상태를 확인한 뒤 다시 시도해 주세요.",
        ) from exc


def _article_similarity_text(article: DonggukMailArticle) -> str:
    return re.sub(r"\s+", " ", f"{article.title} {article.summary or ''}").strip().lower()


def _article_topic_text(article: DonggukMailArticle) -> str:
    text = f"{article.title or ''} {article.summary or ''}".lower()
    text = re.sub(r"\[[^\]]+\]|\([^)]*\)", " ", text)
    text = re.sub(r"\s*-\s*[^-]{2,20}$", " ", text)
    text = re.sub(r"동국대학교|동국대|wise캠퍼스|wise", "동국", text)
    text = re.sub(r"앵커\s*사업|앵커사업|앵커\s*과제|지역성장\s*인재양성체계", "앵커사업", text)
    text = re.sub(r"업무\s*협약|mou|협약\s*체결", "협약", text)
    return re.sub(r"[^가-힣a-z0-9]", "", text)


def _hongbo_topic_bucket(article: DonggukMailArticle) -> str | None:
    text = _dongguk_article_qualification_text(article)
    topic_patterns = [
        ("h안거", r"하안거|30일\s*수행|선.?교\s*겸수"),
        ("lotus_donation", r"로터스관.*(기부|희사|발전기금)|기부.*로터스관|희사.*로터스관"),
        ("iot_degree", r"지능\s*IoT|공동학위"),
        ("meditation_expo", r"서울국제명상엑스포|명상,\s*함께\s*깨어나다|명상엑스포"),
        ("c_forum", r"C포럼"),
        ("anchor_goyang", r"경기앵커|앵커사업단|고양산업진흥원|지역성장\s*인재양성체계"),
        ("oled_operando", r"OLED|오페란도|엑시톤|열화\s*메커니즘"),
        ("dream_workshop", r"Dream\s*Workshop|교사연수"),
        ("academic_member", r"대한민국학술원|학술원\s*신임회원|황훈성"),
        ("mbc_basketball", r"MBC배|대학농구|결선\s*진출"),
        ("education_grant", r"교육교부금|고등교육재정교부금|고등교육\s*투자|사총협"),
        ("ai_basic_education", r"AI\s*기본\s*역량|문.?이과\s*불문\s*AI|신입생\s*AI"),
        ("higher_ed_law", r"고등교육법|시행령|지역인재|입학\s*취소"),
        ("temple_treasure", r"성보|불교문화유산|불교중앙박물관|탑비|벽화|불두|복장물|비파괴조사|보존처리"),
        ("ordination_booth", r"출가상담|출가\s*상담|정광고"),
        ("geunil", r"근일\s*대종사|현봉당\s*근일"),
        ("buddhism_ai", r"AI\s*시대.*불교|불교.*AI|종교.*AI"),
    ]
    for key, pattern in topic_patterns:
        if re.search(pattern, text, re.I):
            return key
    return None


def _article_similarity_tokens(article: DonggukMailArticle) -> set[str]:
    text = _article_similarity_text(article)
    stopwords = {
        "동국대",
        "동국대학교",
        "기사",
        "보도",
        "개최",
        "진행",
        "관련",
        "위해",
        "이번",
        "통해",
        "있는",
        "대한",
        "한다",
        "했다",
        "에서",
        "으로",
        "하고",
    }
    return {
        token
        for token in re.findall(r"[가-힣A-Za-z0-9]{2,}", text)
        if token not in stopwords
    }


def _article_title_topic_tokens(article: DonggukMailArticle) -> set[str]:
    title = (article.title or "").lower()
    title = re.sub(r"앵커\s*사업|앵커사업|앵커\s*과제|지역성장\s*인재양성체계", "앵커사업", title)
    title = re.sub(r"업무\s*협약|mou|협약\s*체결", "협약", title)
    stopwords = {"동국대", "동국대학교", "2026년", "신규", "관련", "기사", "보도", "외"}
    return {
        token
        for token in re.findall(r"[가-힣A-Za-z0-9]{2,}", title)
        if token not in stopwords
    }


def _published_minute_key(article: DonggukMailArticle) -> str:
    value = (article.published_at or "").strip()
    return value[:16] if len(value) >= 16 else value


def _is_similar_dongguk_article(left: DonggukMailArticle, right: DonggukMailArticle) -> bool:
    left_title = re.sub(r"\s+", " ", (left.title or "")).strip().lower()
    right_title = re.sub(r"\s+", " ", (right.title or "")).strip().lower()
    left_bucket = _hongbo_topic_bucket(left)
    right_bucket = _hongbo_topic_bucket(right)
    if left_bucket and left_bucket == right_bucket:
        return True
    if left_title and right_title and SequenceMatcher(None, left_title, right_title).ratio() >= 0.55:
        return True
    if "서울시" in left_title and "서울시" in right_title and "앵커" in left_title and "앵커" in right_title:
        return True
    left_topic = _article_topic_text(left)
    right_topic = _article_topic_text(right)
    if left_topic and right_topic:
        if left_topic in right_topic or right_topic in left_topic:
            return True
        if SequenceMatcher(None, left_topic, right_topic).ratio() >= 0.46:
            return True
    left_title_tokens = _article_title_topic_tokens(left)
    right_title_tokens = _article_title_topic_tokens(right)
    if left_title_tokens and right_title_tokens:
        shared_title_tokens = left_title_tokens & right_title_tokens
        if {"서울시", "앵커사업"} <= shared_title_tokens:
            return True
        if "서울시" in shared_title_tokens and "앵커" in shared_title_tokens:
            return True
        if len(shared_title_tokens) >= 3 and len(shared_title_tokens) / max(1, min(len(left_title_tokens), len(right_title_tokens))) >= 0.5:
            return True
    left_text = _article_similarity_text(left)
    right_text = _article_similarity_text(right)
    if left_text and right_text and SequenceMatcher(None, left_text, right_text).ratio() >= 0.68:
        return True
    left_tokens = _article_similarity_tokens(left)
    right_tokens = _article_similarity_tokens(right)
    if left_tokens and right_tokens:
        overlap = len(left_tokens & right_tokens) / max(1, min(len(left_tokens), len(right_tokens)))
        same_source = (left.source or "").strip() and (left.source or "").strip() == (right.source or "").strip()
        same_minute = _published_minute_key(left) and _published_minute_key(left) == _published_minute_key(right)
        if overlap >= 0.42:
            return True
        if same_source and same_minute and overlap >= 0.22:
            return True
    shared_links = set(_dongguk_article_links(left)) & set(_dongguk_article_links(right))
    return bool(shared_links)


def _merge_dongguk_duplicate(base: DonggukMailArticle, duplicate: DonggukMailArticle) -> DonggukMailArticle:
    base_score = float(base.score or 0)
    duplicate_score = float(duplicate.score or 0)
    representative = duplicate if duplicate_score > base_score else base
    other = base if representative is duplicate else duplicate
    links = list(dict.fromkeys([*_dongguk_article_links(representative), *_dongguk_article_links(other)]))
    data = representative.model_dump()
    data["links"] = links
    data["is_syndicated"] = bool(representative.is_syndicated or other.is_syndicated or len(links) > 1)
    if not data.get("summary"):
        data["summary"] = other.summary
    if not data.get("thumbnail_url"):
        data["thumbnail_url"] = other.thumbnail_url
    return DonggukMailArticle(**data)


def _dedupe_representative_articles(articles: list[DonggukMailArticle]) -> tuple[list[DonggukMailArticle], list[DonggukMailArticle]]:
    kept: list[DonggukMailArticle] = []
    removed: list[DonggukMailArticle] = []
    for article in articles:
        match_index = next((index for index, kept_article in enumerate(kept) if _is_similar_dongguk_article(article, kept_article)), None)
        if match_index is None:
            kept.append(article)
            continue
        merged = _merge_dongguk_duplicate(kept[match_index], article)
        removed_article = article if merged.id == kept[match_index].id else kept[match_index]
        removed_data = removed_article.model_dump()
        removed_data["selection_reason"] = f"같은 주제의 대표 기사 '{merged.title}'로 묶여 메일 대표 목록에서는 제외되었습니다."
        removed_article = DonggukMailArticle(**removed_data)
        removed.append(removed_article)
        kept[match_index] = merged
    return kept, removed


async def _build_dongguk_preview_result(
    *,
    db: AsyncSession,
    current_user: User,
    subject: str,
    mail_date: str | None,
    articles: list[DonggukMailArticle],
    exclude_similar_sent: bool,
    keyword_id: int | None = None,
    priority_criteria: str | None = None,
) -> dict:
    articles, exact_duplicate_excluded = _dedupe_exact_dongguk_articles(articles)
    normalized_priority_criteria = _normalize_dongguk_priority_criteria(priority_criteria)
    normalized_priority_criteria = await PriorityInsightService(db).effective_criteria(
        user_id=current_user.id,
        keyword_id=keyword_id,
        base_criteria=normalized_priority_criteria,
    )
    cache_key = _dongguk_preview_cache_key(
        user_id=current_user.id,
        keyword_id=keyword_id,
        mail_date=mail_date,
        subject=subject,
        articles=articles,
        exclude_similar_sent=exclude_similar_sent,
        priority_criteria=normalized_priority_criteria,
    )
    cached = await _get_cached_dongguk_preview(db, cache_key)
    if cached is not None:
        return cached

    edited_articles, editor_excluded_articles, editor_result = await _run_news_editor_or_fallback(
        current_user=current_user,
        mail_date=mail_date or "",
        subject=subject,
        articles=articles,
        priority_criteria=normalized_priority_criteria,
    )
    sent_titles = await _sent_dongguk_titles(db, current_user.id, before_mail_date=mail_date) if exclude_similar_sent else set()
    included_articles, previously_sent_excluded = _filter_previously_sent_articles(edited_articles, sent_titles)
    included_articles, policy_excluded_articles = _dongguk_mail_section_policy(
        included_articles,
        [*editor_excluded_articles, *exact_duplicate_excluded, *previously_sent_excluded],
        articles,
        mail_date,
    )
    excluded_articles = policy_excluded_articles
    data = {
        "policy_version": DONGGUK_MAIL_POLICY_VERSION,
        "articles": [_dongguk_article_response(article) for article in included_articles],
        "excluded_articles": [_dongguk_article_response(article) for article in excluded_articles],
        "article_count": len(included_articles),
        "excluded_count": len(excluded_articles),
        "editor_used": editor_result is not None,
        "cached": False,
    }
    await _save_dongguk_preview_cache(
        db,
        user_id=current_user.id,
        keyword_id=keyword_id,
        mail_date=mail_date,
        subject=subject,
        cache_key=cache_key,
        data=data,
    )
    return data


@router.post("/dongguk/preview")
async def preview_dongguk_email(
    request: Request,
    body: DonggukPreviewRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_or_dev_user),
):
    mail_date = body.mail_date or _mail_date_from_subject(body.subject)
    if mail_date and not body.force_rebuild:
        draft = await _get_dongguk_mail_draft(
            db,
            user_id=current_user.id,
            keyword_id=body.keyword_id,
            mail_date=mail_date,
        )
        draft_data = _draft_response(draft)
        preview_data = draft_data.get("preview_data")
        if preview_data and preview_data.get("articles") and preview_data.get("editor_used") is not False:
            preview_data["cached"] = True
            preview_data["from_draft"] = True
            return success_response(request=request, data=preview_data)

    data = await _build_dongguk_preview_result(
        db=db,
        current_user=current_user,
        subject=body.subject,
        mail_date=mail_date,
        articles=body.articles,
        exclude_similar_sent=body.exclude_similar_sent,
        keyword_id=body.keyword_id,
        priority_criteria=body.priority_criteria,
    )
    if mail_date:
        selected_articles = [DonggukMailArticle(**item) for item in data.get("articles") or []]
        excluded_articles = [DonggukMailArticle(**item) for item in data.get("excluded_articles") or []]
        await _save_dongguk_mail_draft(
            db,
            user_id=current_user.id,
            keyword_id=body.keyword_id,
            mail_date=mail_date,
            subject=body.subject,
            selected_article_keys=[_dongguk_article_key(article, index) for index, article in enumerate(selected_articles)],
            selected_articles=selected_articles,
            removed_article_keys=[_dongguk_article_key(article, index) for index, article in enumerate(excluded_articles)],
            removed_articles=excluded_articles,
            preview_data=data,
        )
    return success_response(request=request, data=data)


@router.get("/dongguk/draft")
async def get_dongguk_draft(
    request: Request,
    keyword_id: int | None = Query(None),
    mail_date: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_or_dev_user),
):
    draft = await _get_dongguk_mail_draft(
        db,
        user_id=current_user.id,
        keyword_id=keyword_id,
        mail_date=mail_date,
    )
    return success_response(request=request, data=_draft_response(draft))


@router.post("/dongguk/draft")
async def save_dongguk_draft(
    request: Request,
    body: DonggukDraftRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_or_dev_user),
):
    draft = await _save_dongguk_mail_draft(
        db,
        user_id=current_user.id,
        keyword_id=body.keyword_id,
        mail_date=body.mail_date,
        subject=body.subject,
        selected_article_keys=body.selected_article_keys,
        selected_articles=body.selected_articles,
        removed_article_keys=body.removed_article_keys,
        removed_articles=body.removed_articles,
        preview_data=body.preview_data,
        record_feedback=bool(body.feedback_source),
        feedback_source=body.feedback_source or "unknown",
    )
    return success_response(request=request, data=_draft_response(draft))


@router.get("/dongguk/trash")
async def get_dongguk_trash(
    request: Request,
    keyword_id: int | None = Query(None),
    mail_date: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_or_dev_user),
):
    conditions = [DonggukArticleTrash.user_id == current_user.id]
    if keyword_id is not None:
        conditions.append(DonggukArticleTrash.keyword_id == keyword_id)
    result = await db.execute(
        select(DonggukArticleTrash)
        .where(*conditions)
        .order_by(DonggukArticleTrash.created_at.desc())
    )
    items = []
    seen_article_ids: set[int] = set()
    for row in result.scalars().all():
        if row.article_id in seen_article_ids:
            continue
        seen_article_ids.add(row.article_id)
        try:
            article = json.loads(row.article_body or "{}")
        except json.JSONDecodeError:
            article = {"id": row.article_id, "title": "휴지통 기사"}
        items.append(
            {
                "id": row.id,
                "article_id": row.article_id,
                "article": article,
                "mail_date": row.mail_date,
                "trashed_at": _fmt_dt(row.created_at),
            }
        )
    return success_response(request=request, data={"items": items})


@router.post("/dongguk/trash")
async def move_dongguk_article_to_trash(
    request: Request,
    body: DonggukTrashRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_or_dev_user),
):
    article_id = body.article.id
    if article_id is None:
        return success_response(request=request, data={"item": None})
    service = ArticleService(db)
    await service.get_article_detail(user_id=current_user.id, article_id=article_id)
    existing = await db.execute(
        select(DonggukArticleTrash)
        .where(
            DonggukArticleTrash.user_id == current_user.id,
            DonggukArticleTrash.keyword_id == body.keyword_id,
            DonggukArticleTrash.article_id == article_id,
        )
        .limit(1)
    )
    row = existing.scalar_one_or_none()
    if row is None:
        row = DonggukArticleTrash(
            user_id=current_user.id,
            keyword_id=body.keyword_id,
            mail_date=body.mail_date,
            article_id=article_id,
            article_body=json.dumps(body.article.model_dump(), ensure_ascii=False),
        )
        db.add(row)
    else:
        row.mail_date = body.mail_date
        row.article_body = json.dumps(body.article.model_dump(), ensure_ascii=False)
    await PriorityInsightService(db).record_action(
        user_id=current_user.id,
        keyword_id=body.keyword_id,
        mail_date=body.mail_date,
        action_type="trash",
        source_screen="today",
        article=body.article.model_dump(),
        before={"trashed": False},
        after={"trashed": True},
        reason="관리자가 기사를 휴지통으로 이동했습니다.",
    )
    await db.commit()
    await db.refresh(row)
    return success_response(request=request, data={"item": {"id": row.id, "article_id": row.article_id, "article": body.article.model_dump(), "trashed_at": _fmt_dt(row.created_at)}})


@router.post("/dongguk/trash/restore")
async def restore_dongguk_article_from_trash(
    request: Request,
    body: DonggukTrashActionRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_or_dev_user),
):
    existing = await db.scalar(
        select(DonggukArticleTrash)
        .where(
            DonggukArticleTrash.user_id == current_user.id,
            DonggukArticleTrash.keyword_id == body.keyword_id,
            DonggukArticleTrash.article_id == body.article_id,
        )
        .limit(1)
    )
    try:
        article_body = json.loads(existing.article_body or "{}") if existing else {"id": body.article_id}
    except json.JSONDecodeError:
        article_body = {"id": body.article_id}
    await db.execute(
        delete(DonggukArticleTrash).where(
            DonggukArticleTrash.user_id == current_user.id,
            DonggukArticleTrash.keyword_id == body.keyword_id,
            DonggukArticleTrash.article_id == body.article_id,
        )
    )
    await PriorityInsightService(db).record_action(
        user_id=current_user.id,
        keyword_id=body.keyword_id,
        mail_date=body.mail_date,
        action_type="trash_restore",
        source_screen="trash",
        article=article_body,
        before={"trashed": True},
        after={"trashed": False},
        reason="관리자가 휴지통의 기사를 복구했습니다.",
    )
    await db.commit()
    return success_response(request=request, data={"restored": True, "article_id": body.article_id})


@router.post("/dongguk/trash/delete")
async def permanently_delete_dongguk_trashed_article(
    request: Request,
    body: DonggukTrashActionRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_or_dev_user),
):
    existing = await db.execute(
        select(DonggukArticleTrash)
        .where(
            DonggukArticleTrash.user_id == current_user.id,
            DonggukArticleTrash.keyword_id == body.keyword_id,
            DonggukArticleTrash.article_id == body.article_id,
        )
        .limit(1)
    )
    if existing.scalar_one_or_none() is None:
        raise build_error(ErrorCode.NOT_FOUND, "휴지통에 있는 기사만 영구 삭제할 수 있습니다.")
    await db.execute(
        delete(DonggukArticleTrash).where(
            DonggukArticleTrash.user_id == current_user.id,
            DonggukArticleTrash.keyword_id == body.keyword_id,
            DonggukArticleTrash.article_id == body.article_id,
        )
    )
    service = ArticleService(db)
    result = await service.delete_article(user_id=current_user.id, article_id=body.article_id)
    await db.commit()
    return success_response(request=request, data=result.model_dump())


@router.post("/dongguk/email")
async def send_dongguk_email(
    request: Request,
    body: DonggukEmailRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_or_dev_user),
):
    mail_date = body.mail_date or _mail_date_from_subject(body.subject)
    if body.use_current_articles:
        preview_data = {"editor_used": False, "cached": True}
        articles, excluded_articles = _dedupe_exact_dongguk_articles(body.articles)
    else:
        preview_data = await _build_dongguk_preview_result(
            db=db,
            current_user=current_user,
            subject=body.subject,
            mail_date=mail_date or "",
            articles=body.articles,
            exclude_similar_sent=body.exclude_similar_sent,
            keyword_id=body.keyword_id,
            priority_criteria=body.priority_criteria,
        )
        articles = [DonggukMailArticle(**article) for article in preview_data.get("articles") or []]
        excluded_articles = [DonggukMailArticle(**article) for article in preview_data.get("excluded_articles") or []]
    if not articles:
        return success_response(
            request=request,
            data={
                "sent_to": [],
                "article_count": 0,
                "excluded_count": len(excluded_articles),
                "message": "오늘 발송할 신규 기사가 없습니다.",
            },
        )

    effective_subject = f"[테스트] {body.subject}" if body.is_test else body.subject
    html_body = _dongguk_email_html(body.subject, articles)
    text_body = _dongguk_email_text(body.subject, articles)
    delivery_body = _dongguk_delivery_body(body.subject, articles, excluded_articles, mail_date=mail_date)
    hwpx_filename = _dongguk_hwpx_filename(body.subject, mail_date)
    hwpx_bytes = _dongguk_hwpx_bytes(body.subject, articles)
    email_service = EmailService()
    email_service.send_html_email(
        to_emails=[str(email) for email in body.to_emails],
        subject=effective_subject,
        html_body=html_body,
        text_body=text_body,
        from_name="동국대학교 홍보처",
        attachments=[(hwpx_filename, hwpx_bytes, "application/vnd.hancom.hwpx")],
    )
    if not body.is_test:
        sent_at = datetime.now(timezone.utc)
        for email in body.to_emails:
            db.add(
                EmailDelivery(
                    user_id=current_user.id,
                    to_email=str(email),
                    subject=body.subject,
                    body=delivery_body,
                    status="SENT",
                    sent_at=sent_at,
                )
            )
        await db.commit()
    return success_response(
        request=request,
        data={
            "sent_to": body.to_emails,
            "article_count": len(articles),
            "excluded_count": len(excluded_articles),
            "editor_used": preview_data.get("editor_used"),
            "cached": preview_data.get("cached"),
            "is_test": body.is_test,
            "message": "테스트 메일을 발송했습니다." if body.is_test else "홍보처 맞춤 메일을 발송했습니다.",
        },
    )


@router.post("/dongguk/hwp")
async def download_dongguk_hwp(
    body: DonggukHwpRequest,
    current_user: User = Depends(get_current_user_or_dev_user),
):
    del current_user
    hwpx = _dongguk_hwpx_bytes(body.subject, body.articles)
    filename = _dongguk_hwpx_filename(body.subject, body.mail_date)
    encoded_filename = quote(filename)
    return StreamingResponse(
        io.BytesIO(hwpx),
        media_type="application/vnd.hancom.hwpx",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@router.post("/dongguk/link-preview")
async def preview_dongguk_link(
    request: Request,
    body: DonggukLinkPreviewRequest,
    current_user: User = Depends(get_current_user_or_dev_user),
):
    del current_user
    url = body.url.strip()
    if not re.match(r"^https?://", url, re.I):
        return success_response(
            request=request,
            data={
                "title": "추가 기사",
                "source": "직접 입력",
                "summary": "URL을 확인할 수 없습니다. 제목과 요약을 직접 수정해 주세요.",
                "url": url,
                "links": [url] if url else [],
            },
        )

    title = ""
    source = ""
    summary = ""
    thumbnail_url = ""
    try:
        crawl_data = await TransNewsClient().crawl_article(url)
        data = crawl_data.get("data") or {}
        title = data.get("title") or data.get("headline") or ""
        source = data.get("publisher") or data.get("source") or data.get("source_name") or ""
        content = data.get("summary") or data.get("description") or data.get("content") or data.get("text") or ""
        summary = re.sub(r"\s+", " ", str(content)).strip()[:240]
        thumbnail_url = _extract_thumbnail_url(data) or ""
    except Exception as exc:
        print(f"Dongguk link preview fallback url={url}: {exc}")

    if not source:
        source = re.sub(r"^www\.", "", re.sub(r"^https?://", "", url).split("/")[0])
    return success_response(
        request=request,
        data={
            "title": title or f"{source} 추가 기사",
            "source": source or "직접 입력",
            "summary": summary or "추가한 링크입니다. 제목과 요약을 확인 후 수정해 주세요.",
            "url": url,
            "thumbnail_url": thumbnail_url,
            "links": [url],
        },
    )


async def deliver_dongguk_email_for_scheduler(
    *,
    db: AsyncSession,
    current_user: User,
    body: DonggukEmailRequest,
) -> dict:
    mail_date = body.mail_date or _mail_date_from_subject(body.subject)
    preview_data = await _build_dongguk_preview_result(
        db=db,
        current_user=current_user,
        subject=body.subject,
        mail_date=mail_date or "",
        articles=body.articles,
        exclude_similar_sent=body.exclude_similar_sent,
        keyword_id=body.keyword_id,
        priority_criteria=body.priority_criteria,
    )
    articles = [DonggukMailArticle(**article) for article in preview_data.get("articles") or []]
    excluded_articles = [DonggukMailArticle(**article) for article in preview_data.get("excluded_articles") or []]
    if not articles:
        return {
            "sent_to": [],
            "article_count": 0,
            "excluded_count": len(excluded_articles),
            "editor_used": preview_data.get("editor_used"),
            "cached": preview_data.get("cached"),
            "message": "no new articles",
        }

    html_body = _dongguk_email_html(body.subject, articles)
    text_body = _dongguk_email_text(body.subject, articles)
    delivery_body = _dongguk_delivery_body(
        body.subject,
        articles,
        excluded_articles,
        mail_date=mail_date,
        keyword_id=body.keyword_id,
    )
    hwpx_filename = _dongguk_hwpx_filename(body.subject, mail_date)
    hwpx_bytes = _dongguk_hwpx_bytes(body.subject, articles)
    EmailService().send_html_email(
        to_emails=[str(email) for email in body.to_emails],
        subject=body.subject,
        html_body=html_body,
        text_body=text_body,
        from_name="동국대학교 홍보처",
        attachments=[(hwpx_filename, hwpx_bytes, "application/vnd.hancom.hwpx")],
    )
    sent_at = datetime.now(timezone.utc)
    for email in body.to_emails:
        db.add(
            EmailDelivery(
                user_id=current_user.id,
                to_email=str(email),
                subject=body.subject,
                body=delivery_body,
                status="SENT",
                sent_at=sent_at,
            )
        )
    await db.commit()
    return {
        "sent_to": body.to_emails,
        "article_count": len(articles),
        "excluded_count": len(excluded_articles),
        "editor_used": preview_data.get("editor_used"),
        "cached": preview_data.get("cached"),
        "message": "sent",
    }


async def build_dongguk_auto_email_request(
    db: AsyncSession,
    *,
    user_id: int,
    keyword_id: int,
    to_emails: list[str],
    mail_date: str,
) -> DonggukEmailRequest | None:
    keyword = await db.get(Keyword, keyword_id)
    articles = await _dongguk_articles_for_keyword_date(
        db,
        user_id=user_id,
        keyword_id=keyword_id,
        mail_date=mail_date,
    )
    if not articles:
        return None
    return DonggukEmailRequest(
        to_emails=to_emails,
        subject=_dongguk_mail_subject(mail_date),
        keyword_id=keyword_id,
        mail_date=mail_date,
        exclude_similar_sent=True,
        priority_criteria=_normalize_dongguk_priority_criteria(keyword.importance_criteria if keyword else None),
        articles=articles,
    )


async def prebuild_dongguk_mail_drafts_for_scheduler(
    db: AsyncSession,
    *,
    mail_date: str | None = None,
    user_id: int | None = None,
    keyword_ids: list[int] | None = None,
    force_rebuild: bool = False,
) -> dict:
    target_date = mail_date or datetime.now(KST).date().isoformat()
    stmt = (
        select(Keyword, User)
        .join(User, User.id == Keyword.user_id)
        .where(Keyword.is_active.is_(True))
    )
    if user_id is not None:
        stmt = stmt.where(Keyword.user_id == user_id)
    if keyword_ids:
        stmt = stmt.where(Keyword.id.in_(keyword_ids))
    elif user_id is None:
        stmt = stmt.where(Keyword.email_auto_send.is_(True))
    result = await db.execute(stmt)
    rows = [
        row for row in result.all()
        if "동국" in str(getattr(row[0], "keyword_text", "") or getattr(row[0], "keyword", ""))
    ]
    built_count = 0
    skipped_count = 0
    failed_count = 0

    for keyword, user in rows:
        try:
            subject = _dongguk_mail_subject(target_date)
            existing_draft = await _get_dongguk_mail_draft(
                db,
                user_id=user.id,
                keyword_id=keyword.id,
                mail_date=target_date,
            )
            if existing_draft and existing_draft.preview_body and not force_rebuild:
                skipped_count += 1
                continue

            articles = await _dongguk_articles_for_keyword_date(
                db,
                user_id=user.id,
                keyword_id=keyword.id,
                mail_date=target_date,
            )
            if not articles:
                skipped_count += 1
                continue

            if existing_draft and not force_rebuild:
                try:
                    removed_keys = set(json.loads(existing_draft.removed_article_keys or "[]"))
                except json.JSONDecodeError:
                    removed_keys = set()
                try:
                    selected_keys = set(json.loads(existing_draft.selected_article_keys or "[]"))
                except json.JSONDecodeError:
                    selected_keys = set()
                if selected_keys:
                    articles = [
                        article
                        for index, article in enumerate(articles)
                        if _dongguk_article_key(article, index) in selected_keys
                    ]
                if removed_keys:
                    articles = [
                        article
                        for index, article in enumerate(articles)
                        if _dongguk_article_key(article, index) not in removed_keys
                    ]

            if not articles:
                skipped_count += 1
                continue

            preview_data = await _build_dongguk_preview_result(
                db=db,
                current_user=user,
                subject=subject,
                mail_date=target_date,
                articles=articles,
                exclude_similar_sent=True,
                keyword_id=keyword.id,
                priority_criteria=_normalize_dongguk_priority_criteria(keyword.importance_criteria),
            )
            selected_articles = [
                DonggukMailArticle(**item)
                for item in preview_data.get("articles", [])
            ]
            removed_articles = [
                DonggukMailArticle(**item)
                for item in preview_data.get("excluded_articles", [])
            ]
            await _save_dongguk_mail_draft(
                db,
                user_id=user.id,
                keyword_id=keyword.id,
                mail_date=target_date,
                subject=subject,
                selected_article_keys=[
                    _dongguk_article_key(article, index)
                    for index, article in enumerate(selected_articles)
                ],
                selected_articles=selected_articles,
                removed_article_keys=[
                    _dongguk_article_key(article, index)
                    for index, article in enumerate(removed_articles)
                ],
                removed_articles=removed_articles,
                preview_data=preview_data,
            )
            built_count += 1
        except Exception as exc:
            failed_count += 1
            print(f"Dongguk prebuild draft failed keyword_id={getattr(keyword, 'id', None)}: {exc}")

    return {
        "mail_date": target_date,
        "target_count": len(rows),
        "built_count": built_count,
        "skipped_count": skipped_count,
        "failed_count": failed_count,
    }


@router.get("/dongguk/history")
async def dongguk_history(
    request: Request,
    limit: int = Query(30, ge=1, le=100),
    summary_only: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_or_dev_user),
):
    delivery_filters = (
        EmailDelivery.user_id == current_user.id,
        EmailDelivery.status == "SENT",
        EmailDelivery.subject.startswith("오늘의 주요 뉴스 "),
    )
    if summary_only:
        recipient_result = await db.execute(
            select(EmailDelivery.to_email)
            .where(*delivery_filters)
            .order_by(EmailDelivery.sent_at.desc().nullslast(), EmailDelivery.created_at.desc())
            .limit(limit * 10)
        )
        recent_recipients: list[str] = []
        for (email,) in recipient_result.all():
            if email not in recent_recipients:
                recent_recipients.append(email)
        return success_response(
            request=request,
            data={"items": [], "recent_recipients": recent_recipients[:12]},
        )

    distinct_deliveries = (
        select(
            EmailDelivery.id,
            EmailDelivery.subject,
            EmailDelivery.sent_at,
            EmailDelivery.created_at,
            EmailDelivery.body,
        )
        .where(*delivery_filters)
        .distinct(EmailDelivery.subject, EmailDelivery.sent_at)
        .order_by(
            EmailDelivery.subject,
            EmailDelivery.sent_at,
            EmailDelivery.created_at.desc(),
        )
        .subquery()
    )
    result = await db.execute(
        select(distinct_deliveries)
        .order_by(
            distinct_deliveries.c.sent_at.desc().nullslast(),
            distinct_deliveries.c.created_at.desc(),
        )
        .limit(limit)
    )
    deliveries = result.mappings().all()
    subjects = [row["subject"] for row in deliveries]
    recipients_by_subject: dict[str, list[str]] = {}
    recent_recipients: list[str] = []
    if subjects:
        recipient_result = await db.execute(
            select(EmailDelivery.subject, EmailDelivery.to_email)
            .where(EmailDelivery.user_id == current_user.id)
            .where(EmailDelivery.status == "SENT")
            .where(EmailDelivery.subject.in_(subjects))
            .order_by(EmailDelivery.sent_at.desc().nullslast(), EmailDelivery.created_at.desc())
        )
        for subject, email in recipient_result.all():
            recipients = recipients_by_subject.setdefault(subject, [])
            if email not in recipients:
                recipients.append(email)
            if email not in recent_recipients:
                recent_recipients.append(email)

    items: list[dict] = []
    for delivery in deliveries:
        try:
            payload = json.loads(delivery["body"] or "{}")
        except json.JSONDecodeError:
            payload = {}
        if payload.get("type") != "DONGGUK_PR_DAILY":
            continue
        history_articles = [DonggukMailArticle(**article) for article in (payload.get("articles") or [])]
        history_articles, history_duplicates = _dedupe_exact_dongguk_articles(history_articles)
        history_excluded = [DonggukMailArticle(**article) for article in (payload.get("excluded_articles") or [])]
        history_excluded.extend(history_duplicates)
        items.append(
            {
                "id": delivery["id"],
                "subject": delivery["subject"],
                "sent_at": delivery["sent_at"] or delivery["created_at"],
                "recipients": recipients_by_subject.get(delivery["subject"], []),
                "article_count": len(history_articles),
                "excluded_count": len(history_excluded),
                "articles": [{"title": article.title} for article in history_articles[:5]],
            }
        )
    return success_response(
        request=request,
        data={
            "items": items,
            "recent_recipients": recent_recipients[:12],
        },
    )


@router.get("/dongguk/notifications")
async def dongguk_notifications(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_or_dev_user),
):
    items: list[dict] = []

    crawl_result = await db.execute(
        select(CrawlRun)
        .where(CrawlRun.user_id == current_user.id)
        .order_by(CrawlRun.finished_at.desc().nullslast(), CrawlRun.created_at.desc())
        .limit(1)
    )
    crawl_run = crawl_result.scalars().first()
    if crawl_run:
        finished_at = crawl_run.finished_at or crawl_run.created_at
        status_label = "완료" if crawl_run.status == "COMPLETED" else "진행 중"
        items.append(
            {
                "id": f"crawl-{crawl_run.id}",
                "type": "crawl",
                "title": f"기사 수집 {status_label}",
                "message": f"{crawl_run.article_count or 0}건 수집됨",
                "created_at": finished_at.isoformat() if finished_at else None,
                "read": False,
            }
        )

    draft_result = await db.execute(
        select(DonggukMailDraft)
        .where(DonggukMailDraft.user_id == current_user.id)
        .where(DonggukMailDraft.preview_body.is_not(None))
        .order_by(DonggukMailDraft.updated_at.desc(), DonggukMailDraft.created_at.desc())
        .limit(1)
    )
    draft = draft_result.scalars().first()
    if draft:
        try:
            preview_data = json.loads(draft.preview_body or "{}")
        except json.JSONDecodeError:
            preview_data = {}
        article_count = len(preview_data.get("articles") or [])
        excluded_count = len(preview_data.get("excluded_articles") or [])
        created_at = draft.updated_at or draft.created_at
        items.append(
            {
                "id": f"dongguk-dify-{draft.id}",
                "type": "dify",
                "title": "Dify 우선순위 판정 완료",
                "message": f"{draft.mail_date} 대표 기사 {article_count}건, 제외 {excluded_count}건",
                "created_at": created_at.isoformat() if created_at else None,
                "read": False,
            }
        )

    items.sort(key=lambda item: item.get("created_at") or "", reverse=True)
    return success_response(request=request, data={"items": items[:6]})


@router.get("/daily")
async def download_daily_report(
    keyword_id: int | None = Query(None, description="?뱀젙 ?ㅼ썙???꾪꽣"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_or_dev_user),
):
    keyword_name = await _keyword_name(db, current_user, keyword_id)
    rows = await _report_rows(db, current_user, keyword_id)
    social_rows = await _get_social_report_rows(db, current_user, keyword_id=keyword_id)

    wb = _build_excel(rows, keyword_name=keyword_name, social_rows=social_rows)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename = f"daily_report_{today_str}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/email")
async def send_daily_report_email(
    request: Request,
    body: EmailReportRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_or_dev_user),
):
    keyword_name = body.keyword_name or await _keyword_name(db, current_user, body.keyword_id)
    rows = await _report_rows(db, current_user, body.keyword_id)
    social_rows = await _get_social_report_rows(db, current_user, keyword_id=body.keyword_id)

    wb = _build_excel(rows, keyword_name=keyword_name, social_rows=social_rows)
    stream = io.BytesIO()
    wb.save(stream)
    excel_bytes = stream.getvalue()

    context = _report_context(rows, keyword_name, social_rows)
    email_service = EmailService()
    email_service.send_daily_report(
        to_emails=body.to_emails,
        excel_bytes=excel_bytes,
        keyword_name=keyword_name,
        html_body=_build_email_html(context),
        text_body=_build_email_text(context),
    )

    return success_response(
        request=request,
        data={
            "sent_to": body.to_emails,
            "article_count": len(rows),
            "message": f"{len(body.to_emails)}紐낆뿉寃??곗씪由?由ы룷?몃? 諛쒖넚?덉뒿?덈떎.",
        },
    )
