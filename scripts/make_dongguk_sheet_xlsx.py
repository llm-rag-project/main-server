import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_PATH = Path("/code/exports/dongguk_sheet_rows_latest.json")
OUTPUT_PATH = Path("/code/exports/dongguk_sheet_rows_latest.xlsx")
HEADERS = ["기준일", "포함여부", "순번", "섹션", "수집 출처", "제목", "발행일시", "URL", "수집풀"]


def normalize_section(row):
    section = row.get("섹션") or "미분류"
    title = row.get("제목") or ""
    pool = row.get("수집풀") or ""
    if section == "미분류" and ("동국" in title or pool in {"dongguk_core", "dongguk_media", "dongguk_keyword"}):
        return "동국대 [법인/건학위]"
    return section


def main():
    data = json.loads(INPUT_PATH.read_text(encoding="utf-8"))
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(fill_type="solid", fgColor="DDEBFF")
    included_fill = PatternFill(fill_type="solid", fgColor="EAF7EA")
    excluded_fill = PatternFill(fill_type="solid", fgColor="F7F7F7")

    for mail_date, info in data["dates"].items():
        ws = wb.create_sheet(mail_date)
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in info["rows"]:
            normalized = dict(row)
            normalized["섹션"] = normalize_section(normalized)
            ws.append([normalized.get(header) or "" for header in HEADERS])

        for row in ws.iter_rows(min_row=2):
            fill = included_fill if row[1].value == "메일 포함" else excluded_fill
            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        widths = [12, 12, 8, 20, 24, 80, 24, 70, 18]
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
